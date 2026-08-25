"""Streaming, resumable validation for large FMEDA workbooks."""

from __future__ import annotations

import hashlib
import json
import math
import re
import shutil
import time
import zipfile
from datetime import datetime, timezone
from itertools import zip_longest
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

try:
    from lxml import etree as LET
except ImportError:  # pragma: no cover - exercised only in minimal environments
    LET = None

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _iter_xml(source: Any, tags: tuple[str, ...] | None = None):
    if LET is not None:
        kwargs = {"events": ("end",), "huge_tree": True}
        if tags:
            kwargs["tag"] = tags
        return LET.iterparse(source, **kwargs)
    return ET.iterparse(source, events=("end",))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _is_error(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("#")


def _value_key(value: Any) -> str:
    return json.dumps(_jsonable(value), ensure_ascii=False, sort_keys=True)


def _values_equal(left: Any, right: Any) -> bool:
    if isinstance(left, (int, float)) and not isinstance(left, bool) and isinstance(right, (int, float)) and not isinstance(right, bool):
        return math.isclose(float(left), float(right), rel_tol=1e-12, abs_tol=1e-12)
    return _value_key(left) == _value_key(right)


def _canonical_formula(formula: str | None) -> str | None:
    if formula is None:
        return None
    result: list[str] = []
    in_string = False
    for char in formula:
        if char == '"':
            in_string = not in_string
            result.append(char)
        elif char.isspace() and not in_string:
            continue
        elif not in_string:
            result.append(char.upper())
        else:
            result.append(char)
    normalized = "".join(result)
    normalized = re.sub(r"\b(FALSE|TRUE)\(\)", r"\1", normalized)
    return normalized


def _read_zip_sheet_paths(path: Path) -> dict[str, str]:
    with zipfile.ZipFile(path, "r") as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets: dict[str, str] = {}
        for rel in relationships:
            if _local(rel.tag) != "Relationship":
                continue
            target = rel.attrib.get("Target", "").replace("\\", "/").lstrip("/")
            if not target.startswith("xl/"):
                target = f"xl/{target}"
            targets[rel.attrib.get("Id", "")] = target
        result: dict[str, str] = {}
        for sheet in workbook.iter(f"{{{MAIN_NS}}}sheet"):
            name = sheet.attrib.get("name", "")
            rel_id = sheet.attrib.get(f"{{{REL_NS}}}id", "")
            if name and rel_id in targets:
                result[name] = targets[rel_id]
        return result


_METADATA_TAGS = tuple(
    f"{{{MAIN_NS}}}{kind}"
    for kind in (
        "dimension",
        "mergeCell",
        "dataValidation",
        "conditionalFormatting",
        "tablePart",
        "pane",
    )
)


def _worksheet_metadata_one(path: Path, xml_path: str) -> dict[str, Any]:
    """Read one worksheet's structural metadata without materializing its cells."""
    dimension = None
    merges: list[str] = []
    data_validation_count = 0
    conditional_formatting_count = 0
    table_part_count = 0
    freeze_panes = None
    with zipfile.ZipFile(path, "r") as archive:
        with archive.open(xml_path, "r") as stream:
            for _, node in _iter_xml(stream, _METADATA_TAGS):
                kind = _local(node.tag)
                if kind == "dimension" and dimension is None:
                    dimension = node.attrib.get("ref")
                elif kind == "mergeCell" and node.attrib.get("ref"):
                    merges.append(node.attrib["ref"])
                elif kind == "dataValidation":
                    data_validation_count += 1
                elif kind == "conditionalFormatting":
                    conditional_formatting_count += 1
                elif kind == "tablePart":
                    table_part_count += 1
                elif kind == "pane" and node.attrib.get("topLeftCell"):
                    freeze_panes = node.attrib["topLeftCell"]
                node.clear()
    return {
        "dimension_ref": dimension,
        "merged_ranges": sorted(merges),
        "merged_range_count": len(merges),
        "data_validation_count": data_validation_count,
        "conditional_formatting_count": conditional_formatting_count,
        "table_part_count": table_part_count,
        "freeze_panes": freeze_panes,
    }


def _cell_record(
    formula_cell: Any,
    cached_cell: Any,
    coordinate: str,
) -> dict[str, Any] | None:
    raw = formula_cell.value if formula_cell is not None else None
    cached = cached_cell.value if cached_cell is not None else None
    if raw is None and cached is None:
        return None
    formula_raw = raw if getattr(formula_cell, "data_type", None) == "f" or (
        isinstance(raw, str) and raw.startswith("=")
    ) else None
    raw_value = None if formula_raw is not None else raw
    if raw is None and cached is None and formula_raw is None:
        return None
    error_state = (
        "error"
        if getattr(formula_cell, "data_type", None) == "e"
        or _is_error(raw_value)
        or _is_error(cached)
        else "ok"
    )
    return {
        "cell": coordinate,
        "formula_raw": _jsonable(formula_raw),
        "raw_value": _jsonable(raw_value),
        "cached_value": _jsonable(cached),
        "error_state": error_state,
    }


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    values: list[str] = []
    with archive.open("xl/sharedStrings.xml", "r") as stream:
        for _, node in _iter_xml(stream, (f"{{{MAIN_NS}}}si",)):
            if _local(node.tag) == "si":
                values.append("".join(child.text or "" for child in node.iter() if _local(child.tag) == "t"))
                node.clear()
    return values


def _xml_value(cell: ET.Element, shared_strings: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(child.text or "" for child in cell.iter() if _local(child.tag) == "t")
    value_node = next((child for child in cell if _local(child.tag) == "v"), None)
    raw = "" if value_node is None or value_node.text is None else value_node.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return raw
    if cell_type == "b":
        return raw == "1"
    if cell_type in {"e", "str"}:
        return raw
    if raw == "":
        return None
    try:
        if "." in raw or "e" in raw.lower():
            return float(raw)
        return int(raw)
    except ValueError:
        return raw


def _xml_cell_record(
    cell: ET.Element,
    shared_strings: list[str],
    shared_formulas: dict[str, tuple[str, str]],
) -> dict[str, Any]:
    formula_node = next((child for child in cell if _local(child.tag) == "f"), None)
    formula_raw = None
    formula_signature = None
    if formula_node is not None:
        text = formula_node.text or ""
        formula_type = formula_node.attrib.get("t")
        formula_si = formula_node.attrib.get("si")
        if formula_type == "shared" and formula_si:
            if text:
                formula_raw = f"={text}"
                shared_formulas[formula_si] = (cell.attrib.get("r", ""), formula_raw)
            else:
                anchor = shared_formulas.get(formula_si)
                if anchor:
                    try:
                        formula_raw = Translator(anchor[1], origin=anchor[0]).translate_formula(
                            cell.attrib.get("r", "")
                        )
                    except (KeyError, ValueError):
                        formula_raw = anchor[1]
        else:
            formula_raw = f"={text}" if text else None
        # Compare semantic formula text, not whether the producer encoded it as shared/array.
        formula_signature = _canonical_formula(formula_raw)
    cached_value = _xml_value(cell, shared_strings)
    raw_value = None if formula_node is not None else cached_value
    error_state = "error" if cell.attrib.get("t") == "e" or _is_error(raw_value) or _is_error(cached_value) else "ok"
    return {
        "cell": cell.attrib.get("r", ""),
        "formula_raw": formula_raw,
            "formula_signature": formula_signature,
        "raw_value": _jsonable(raw_value),
        "cached_value": _jsonable(cached_value),
        "error_state": error_state,
    }


def _xml_sheet_cells(path: Path, xml_path: str):
    with zipfile.ZipFile(path, "r") as archive:
        shared_strings = _read_shared_strings(archive)
        shared_formulas: dict[str, tuple[str, str]] = {}
        with archive.open(xml_path, "r") as stream:
            for _, node in _iter_xml(stream, (f"{{{MAIN_NS}}}c",)):
                if _local(node.tag) == "c":
                    record = _xml_cell_record(node, shared_strings, shared_formulas)
                    if record["cell"] and (
                        record["formula_signature"] is not None
                        or record["raw_value"] is not None
                        or record["cached_value"] is not None
                    ):
                        yield record
                    node.clear()


def _compare_xml_cell_pair(
    key: str,
    before: dict[str, Any] | None,
    after: dict[str, Any] | None,
    allowed: dict[str, dict[str, Any]],
    differences: list[dict[str, Any]],
) -> None:
    if before is None or after is None:
        differences.append({
            "key": key,
            "kind": "cell_presence",
            "blocking": True,
            "status": "unexpected",
            "before": before,
            "after": after,
        })
        return
    if before["formula_signature"] != after["formula_signature"]:
        differences.append({
            "key": key,
            "kind": "formula_raw",
            "blocking": True,
            "status": "unexpected",
            "before": before["formula_raw"] or before["formula_signature"],
            "after": after["formula_raw"] or after["formula_signature"],
        })
        return
    if before["error_state"] != after["error_state"]:
        differences.append({
            "key": key,
            "kind": "error_state",
            "blocking": True,
            "status": "unexpected",
            "before": before["error_state"],
            "after": after["error_state"],
        })
    if before["formula_signature"] is not None:
        if not _values_equal(before["cached_value"], after["cached_value"]):
            differences.append({
                "key": key,
                "kind": "cached_value",
                "blocking": False,
                "status": "warning",
                "before": before["cached_value"],
                "after": after["cached_value"],
            })
        return
    if _values_equal(before["raw_value"], after["raw_value"]):
        return
    expected = allowed.get(key)
    if expected and _values_equal(expected.get("old_value"), before["raw_value"]) and _values_equal(
        expected.get("new_value"), after["raw_value"]
    ):
        differences.append({
            "key": key,
            "kind": "allowed_input",
            "blocking": False,
            "status": "allowed",
            "before": before["raw_value"],
            "after": after["raw_value"],
        })
    else:
        differences.append({
            "key": key,
            "kind": "unexpected_value",
            "blocking": True,
            "status": "unexpected",
            "before": before["raw_value"],
            "after": after["raw_value"],
        })


def _compare_sheet_xml(
    base_path: Path,
    base_xml_path: str,
    target_path: Path,
    target_xml_path: str,
    sheet_name: str,
    allowed: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    differences: list[dict[str, Any]] = []
    base_cells = _xml_sheet_cells(base_path, base_xml_path)
    target_cells = _xml_sheet_cells(target_path, target_xml_path)
    before = next(base_cells, None)
    after = next(target_cells, None)
    while before is not None or after is not None:
        before_key = before["cell"] if before is not None else None
        after_key = after["cell"] if after is not None else None
        if before_key == after_key:
            key = f"{sheet_name}!{before_key}"
            _compare_xml_cell_pair(key, before, after, allowed, differences)
            before = next(base_cells, None)
            after = next(target_cells, None)
        elif after_key is None or (before_key is not None and before_key < after_key):
            _compare_xml_cell_pair(
                f"{sheet_name}!{before_key}", before, None, allowed, differences
            )
            before = next(base_cells, None)
        else:
            _compare_xml_cell_pair(
                f"{sheet_name}!{after_key}", None, after, allowed, differences
            )
            after = next(target_cells, None)
    return differences


def _compare_sheet(
    sheet_name: str,
    base_formula_sheet: Any,
    target_formula_sheet: Any,
    base_cached_sheet: Any,
    target_cached_sheet: Any,
    allowed: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    differences: list[dict[str, Any]] = []
    sentinel = object()
    rows = zip_longest(
        base_formula_sheet.iter_rows(),
        target_formula_sheet.iter_rows(),
        base_cached_sheet.iter_rows(),
        target_cached_sheet.iter_rows(),
        fillvalue=(),
    )
    for row_number, (base_row, target_row, base_cached_row, target_cached_row) in enumerate(rows, start=1):
        max_len = max(len(base_row), len(target_row), len(base_cached_row), len(target_cached_row))
        for index in range(max_len):
            base_cell = base_row[index] if index < len(base_row) else None
            target_cell = target_row[index] if index < len(target_row) else None
            base_cached = base_cached_row[index] if index < len(base_cached_row) else None
            target_cached = target_cached_row[index] if index < len(target_cached_row) else None
            coordinate = f"{get_column_letter(index + 1)}{row_number}"
            before = _cell_record(base_cell, base_cached, coordinate)
            after = _cell_record(target_cell, target_cached, coordinate)
            if before is None and after is None:
                continue
            key = f"{sheet_name}!{(before or after)['cell']}"
            if before is None or after is None:
                differences.append(
                    {
                        "key": key,
                        "kind": "cell_presence",
                        "blocking": True,
                        "status": "unexpected",
                        "before": before,
                        "after": after,
                    }
                )
                continue
            if before["formula_raw"] != after["formula_raw"]:
                differences.append(
                    {
                        "key": key,
                        "kind": "formula_raw",
                        "blocking": True,
                        "status": "unexpected",
                        "before": before["formula_raw"],
                        "after": after["formula_raw"],
                    }
                )
                continue
            if before["error_state"] != after["error_state"]:
                differences.append(
                    {
                        "key": key,
                        "kind": "error_state",
                        "blocking": True,
                        "status": "unexpected",
                        "before": before["error_state"],
                        "after": after["error_state"],
                    }
                )
            if before["formula_raw"] is not None:
                if not _values_equal(before["cached_value"], after["cached_value"]):
                    differences.append(
                        {
                            "key": key,
                            "kind": "cached_value",
                            "blocking": False,
                            "status": "warning",
                            "before": before["cached_value"],
                            "after": after["cached_value"],
                        }
                    )
                continue
            if _values_equal(before["raw_value"], after["raw_value"]):
                continue
            expected = allowed.get(key)
            if expected and _value_key(expected.get("old_value")) == _value_key(before["raw_value"]) and _value_key(
                expected.get("new_value")
            ) == _value_key(after["raw_value"]):
                differences.append(
                    {
                        "key": key,
                        "kind": "allowed_input",
                        "blocking": False,
                        "status": "allowed",
                        "before": before["raw_value"],
                        "after": after["raw_value"],
                    }
                )
            else:
                differences.append(
                    {
                        "key": key,
                        "kind": "unexpected_value",
                        "blocking": True,
                        "status": "unexpected",
                        "before": before["raw_value"],
                        "after": after["raw_value"],
                    }
                )
    return differences


def _patch_changes(patch_manifest: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for change in (patch_manifest or {}).get("changes", []):
        if not isinstance(change, dict):
            continue
        sheet = str(change.get("sheet") or "")
        cell = str(change.get("cell") or "").upper().replace("$", "")
        if sheet and cell:
            result[f"{sheet}!{cell}"] = change
    return result


def _atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temp.replace(path)


class LargeFmedaValidator:
    """Validate large workbook revisions with per-sheet JSONL checkpoints."""

    def __init__(
        self,
        base: str | Path,
        target: str | Path,
        output_dir: str | Path,
        patch_manifest: dict[str, Any] | None = None,
    ):
        self.base = Path(base).expanduser().resolve()
        self.target = Path(target).expanduser().resolve()
        self.output_dir = Path(output_dir).expanduser().resolve()
        self.patch_manifest = patch_manifest or {}
        if not self.base.is_file():
            raise FileNotFoundError(f"base workbook not found: {self.base}")
        if not self.target.is_file():
            raise FileNotFoundError(f"target workbook not found: {self.target}")

    @staticmethod
    def sha256(path: str | Path) -> str:
        return _sha256(Path(path))

    def _load_checkpoint(self, base_hash: str, target_hash: str) -> dict[str, Any]:
        checkpoint_path = self.output_dir / "checkpoint.json"
        if not checkpoint_path.is_file():
            return {
                "schema_version": "fmeda-large-validation-checkpoint-v1",
                "base_sha256": base_hash,
                "target_sha256": target_hash,
                "completed_sheet_indices": [],
                "resumed_sheet_indices": [],
                "status": "new",
            }
        checkpoint = json.loads(checkpoint_path.read_text(encoding="utf-8"))
        if checkpoint.get("base_sha256") != base_hash or checkpoint.get("target_sha256") != target_hash:
            return {
                "schema_version": "fmeda-large-validation-checkpoint-v1",
                "base_sha256": base_hash,
                "target_sha256": target_hash,
                "completed_sheet_indices": [],
                "resumed_sheet_indices": [],
                "status": "reset_due_to_revision_change",
            }
        checkpoint.setdefault("completed_sheet_indices", [])
        checkpoint.setdefault("resumed_sheet_indices", [])
        return checkpoint

    def _iter_chunk_differences(self, completed: set[int], sheet_paths: dict[str, str]):
        paths = [
            self.output_dir / "chunks" / f"{index:02d}_{self._safe_sheet(sheet_paths, index)}.jsonl"
            for index in sorted(completed)
        ]
        workbook_chunk = self.output_dir / "chunks" / "00_workbook.jsonl"
        if workbook_chunk.is_file():
            paths.append(workbook_chunk)
        for path in paths:
            if not path.is_file():
                continue
            with path.open("r", encoding="utf-8") as handle:
                for line in handle:
                    if line.strip():
                        yield json.loads(line)

    def _summarize_chunk_differences(
        self,
        completed: set[int],
        sheet_paths: dict[str, str],
    ) -> dict[str, Any]:
        samples: list[dict[str, Any]] = []
        difference_count = 0
        blocking_count = 0
        warning_count = 0
        allowed_count = 0
        for item in self._iter_chunk_differences(completed, sheet_paths):
            difference_count += 1
            if item.get("blocking"):
                blocking_count += 1
            if item.get("status") == "warning":
                warning_count += 1
            if item.get("status") == "allowed":
                allowed_count += 1
            if len(samples) < 200:
                samples.append(item)
        return {
            "samples": samples,
            "difference_count": difference_count,
            "blocking_count": blocking_count,
            "warning_count": warning_count,
            "allowed_count": allowed_count,
        }

    @staticmethod
    def _safe_sheet(sheet_paths: dict[str, str], index: int) -> str:
        names = list(sheet_paths)
        if 1 <= index <= len(names):
            return "".join(char if char.isalnum() or char in "._-" else "_" for char in names[index - 1])[:80] or "Sheet"
        return "Sheet"

    def _write_sheet_chunk(self, index: int, sheet_name: str, differences: list[dict[str, Any]]) -> None:
        chunk_dir = self.output_dir / "chunks"
        chunk_dir.mkdir(parents=True, exist_ok=True)
        safe_name = "".join(char if char.isalnum() or char in "._-" else "_" for char in sheet_name)[:80] or "Sheet"
        path = chunk_dir / f"{index:02d}_{safe_name}.jsonl"
        with path.open("w", encoding="utf-8") as handle:
            for item in differences:
                handle.write(json.dumps(item, ensure_ascii=False) + "\n")

    def run(self, max_sheets: int | None = None) -> dict[str, Any]:
        started = time.monotonic()
        self.output_dir.mkdir(parents=True, exist_ok=True)
        base_hash = _sha256(self.base)
        target_hash = _sha256(self.target)
        checkpoint = self._load_checkpoint(base_hash, target_hash)
        completed: set[int] = {int(index) for index in checkpoint.get("completed_sheet_indices", [])}

        base_paths = _read_zip_sheet_paths(self.base)
        target_paths = _read_zip_sheet_paths(self.target)
        base_sheet_names = list(base_paths)
        target_sheet_names = list(target_paths)
        metadata_path = self.output_dir / "worksheet_metadata.json"
        if metadata_path.is_file():
            metadata_cache = json.loads(metadata_path.read_text(encoding="utf-8"))
        else:
            metadata_cache = {"schema_version": "fmeda-worksheet-metadata-v1", "base": {}, "target": {}}
        base_meta = metadata_cache.setdefault("base", {})
        target_meta = metadata_cache.setdefault("target", {})
        allowed = _patch_changes(self.patch_manifest)

        workbook_metadata_differences: list[dict[str, Any]] = []
        if base_sheet_names != target_sheet_names:
            workbook_metadata_differences.append(
                {
                    "key": "workbook",
                    "kind": "sheet_order_or_presence",
                    "blocking": True,
                    "status": "unexpected",
                    "before": base_sheet_names,
                    "after": target_sheet_names,
                }
            )
        if workbook_metadata_differences:
            self.output_dir.joinpath("chunks").mkdir(parents=True, exist_ok=True)
            with (self.output_dir / "chunks" / "00_workbook.jsonl").open("w", encoding="utf-8") as handle:
                for item in workbook_metadata_differences:
                    handle.write(json.dumps(item, ensure_ascii=False) + "\n")

        process_count = 0
        resumed: list[int] = []
        new_completed = set(completed)
        for index, sheet_name in enumerate(target_sheet_names, start=1):
            if sheet_name not in base_paths:
                continue
            if sheet_name not in base_meta:
                base_meta[sheet_name] = _worksheet_metadata_one(self.base, base_paths[sheet_name])
            if sheet_name not in target_meta:
                target_meta[sheet_name] = _worksheet_metadata_one(self.target, target_paths[sheet_name])
            _atomic_write_json(metadata_path, metadata_cache)
            if index in completed:
                resumed.append(index)
                continue
            if max_sheets is not None and process_count >= max_sheets:
                break
            sheet_differences = _compare_sheet_xml(
                self.base,
                base_paths[sheet_name],
                self.target,
                target_paths[sheet_name],
                sheet_name,
                allowed,
            )
            base_struct = base_meta.get(sheet_name, {})
            target_struct = target_meta.get(sheet_name, {})
            if base_struct.get("merged_ranges") != target_struct.get("merged_ranges"):
                sheet_differences.append(
                    {
                        "key": f"{sheet_name}!__metadata__",
                        "kind": "merged_cells",
                        "blocking": True,
                        "status": "unexpected",
                        "before": base_struct.get("merged_ranges", []),
                        "after": target_struct.get("merged_ranges", []),
                    }
                )
            metadata_fields = (
                "dimension_ref",
                "data_validation_count",
                "conditional_formatting_count",
                "table_part_count",
                "freeze_panes",
            )
            before_metadata = {field: base_struct.get(field) for field in metadata_fields}
            after_metadata = {field: target_struct.get(field) for field in metadata_fields}
            if before_metadata != after_metadata:
                sheet_differences.append(
                    {
                        "key": f"{sheet_name}!__metadata__",
                        "kind": "worksheet_metadata",
                        "blocking": True,
                        "status": "unexpected",
                        "before": before_metadata,
                        "after": after_metadata,
                    }
                )
            self._write_sheet_chunk(index, sheet_name, sheet_differences)
            new_completed.add(index)
            process_count += 1
            checkpoint.update(
                {
                    "completed_sheet_indices": sorted(new_completed),
                    "last_completed_sheet_index": index,
                    "updated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
                    "status": "running",
                }
            )
            _atomic_write_json(self.output_dir / "checkpoint.json", checkpoint)

        all_sheets_completed = set(range(1, len(target_sheet_names) + 1)).issubset(new_completed)
        difference_summary = self._summarize_chunk_differences(new_completed, target_paths)
        if not all_sheets_completed:
            status = "INCOMPLETE"
        elif difference_summary["blocking_count"]:
            status = "FAIL"
        elif difference_summary["warning_count"]:
            status = "PASS_WITH_WARNINGS"
        else:
            status = "PASS"
        checkpoint.update(
            {
                "completed_sheet_indices": sorted(new_completed),
                "resumed_sheet_indices": resumed,
                "status": "completed" if all_sheets_completed else "incomplete",
                "updated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            }
        )
        _atomic_write_json(self.output_dir / "checkpoint.json", checkpoint)
        sheet_summaries = []
        for index, sheet_name in enumerate(target_sheet_names, start=1):
            if index not in new_completed:
                continue
            base_struct = base_meta.get(sheet_name, {})
            target_struct = target_meta.get(sheet_name, {})
            sheet_summaries.append(
                {
                    "index": index,
                    "name": sheet_name,
                    "merged_range_count": target_struct.get("merged_range_count", 0),
                    "merged_cells_preserved": base_struct.get("merged_ranges", []) == target_struct.get("merged_ranges", []),
                    "data_validation_count": target_struct.get("data_validation_count", 0),
                    "conditional_formatting_count": target_struct.get("conditional_formatting_count", 0),
                    "table_part_count": target_struct.get("table_part_count", 0),
                }
            )
        report = {
            "schema_version": "fmeda-large-validation-report-v1",
            "status": status,
            "base_file": self.base.name,
            "target_file": self.target.name,
            "base_sha256": base_hash,
            "target_sha256": target_hash,
            "sheet_count": len(target_sheet_names),
            "completed_sheet_count": len(new_completed),
            "summary": {
                "allowed_input_changes": difference_summary["allowed_count"],
                "blocking_changes": difference_summary["blocking_count"],
                "warnings": difference_summary["warning_count"],
                "difference_count": difference_summary["difference_count"],
            },
            "checkpoint": checkpoint,
            "sheet_summaries": sheet_summaries,
            "samples": difference_summary["samples"],
            "duration_seconds": round(time.monotonic() - started, 3),
        }
        _atomic_write_json(self.output_dir / "validation.json", report)
        markdown = self.render_markdown(report)
        (self.output_dir / "validation.md").write_text(markdown, encoding="utf-8")
        return report

    @staticmethod
    def render_markdown(report: dict[str, Any]) -> str:
        lines = [
            "# Large FMEDA Revision Validation Report",
            "",
            f"**Status**: `{report['status']}`  ",
            f"**Base**: `{report['base_file']}`  ",
            f"**Target**: `{report['target_file']}`  ",
            f"**Sheets**: `{report['completed_sheet_count']}/{report['sheet_count']}` completed  ",
            f"**Base SHA-256**: `{report['base_sha256']}`  ",
            f"**Target SHA-256**: `{report['target_sha256']}`  ",
            "",
            "## Summary",
            "",
            "| Metric | Count |",
            "|---|---:|",
            f"| Allowed input changes | {report['summary']['allowed_input_changes']} |",
            f"| Blocking changes | {report['summary']['blocking_changes']} |",
            f"| Warnings | {report['summary']['warnings']} |",
            f"| Total differences | {report['summary']['difference_count']} |",
            "",
            "## Differences",
            "",
            "| Location | Kind | Before | After | Result |",
            "|---|---|---|---|---|",
        ]
        if report["samples"]:
            for item in report["samples"]:
                before = item.get("before")
                after = item.get("after")
                lines.append(
                    f"| `{item['key']}` | `{item['kind']}` | "
                    f"`{str(before).replace('|', '\\|')}` | "
                    f"`{str(after).replace('|', '\\|')}` | `{item['status']}` |"
                )
        else:
            lines.append("| — | none | — | — | `unchanged` |")
        lines.extend(
            [
                "",
                "## Merged cells and metadata",
                "",
                "| Sheet | Merged ranges | Preserved | Data validations | Conditional formatting | Tables |",
                "|---|---:|---|---:|---:|---:|",
            ]
        )
        for sheet in report["sheet_summaries"]:
            lines.append(
                f"| `{sheet['name']}` | {sheet['merged_range_count']} | "
                f"`{sheet['merged_cells_preserved']}` | {sheet['data_validation_count']} | "
                f"{sheet['conditional_formatting_count']} | {sheet['table_part_count']} |"
            )
        lines.extend(
            [
                "",
                "> 此報告使用 Excel-compatible 公式原文與 cached value 進行差異驗證；它不把 cached value 變化誤認為已完成獨立重算。",
                "",
            ]
        )
        return "\n".join(lines)
