#!/usr/bin/env python3
"""
image_extractor.py - Robust image extraction from Excel files.

Handles various image formats and storage methods in Excel files.
"""

import logging
import io
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from PIL import Image

import openpyxl
from openpyxl.drawing.image import Image as XLImage

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ImageExtractor:
    """Extract images from Excel files robustly."""
    
    # Supported image formats
    SUPPORTED_FORMATS = {
        'jpeg': '.jpg',
        'jpg': '.jpg',
        'png': '.png',
        'gif': '.gif',
        'bmp': '.bmp',
        'tiff': '.tiff',
        'wmf': '.wmf',
        'emf': '.emf'
    }
    
    def __init__(self, output_dir: Path = Path("extracted_images")):
        """
        Initialize image extractor.
        
        Args:
            output_dir: Directory to save extracted images
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Initialized ImageExtractor with output dir: {self.output_dir}")
    
    def extract_from_workbook(self, xlsx_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        Extract images from all sheets in workbook.
        
        Args:
            xlsx_path: Path to Excel file
            
        Returns:
            Dictionary mapping sheet names to list of extracted image info
        """
        wb = openpyxl.load_workbook(xlsx_path)
        results = {}
        
        logger.info(f"Extracting images from {xlsx_path}")
        logger.info(f"Total sheets: {len(wb.sheetnames)}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            images = self.extract_from_sheet(ws, sheet_name)
            if images:
                results[sheet_name] = images
                logger.info(f"  Sheet '{sheet_name}': {len(images)} images extracted")
        
        return results
    
    def extract_from_sheet(self, ws, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Extract images from a single sheet.
        
        Args:
            ws: Worksheet object
            sheet_name: Name of the sheet
            
        Returns:
            List of extracted image information
        """
        images = []
        
        if not hasattr(ws, '_images') or not ws._images:
            return images
        
        for idx, img in enumerate(ws._images, 1):
            try:
                image_info = self._extract_single_image(img, sheet_name, idx)
                if image_info:
                    images.append(image_info)
            except Exception as e:
                logger.error(f"Error extracting image {idx} from {sheet_name}: {e}")
        
        return images
    
    def _extract_single_image(self, img, sheet_name: str, idx: int) -> Optional[Dict[str, Any]]:
        """
        Extract a single image from worksheet.
        
        Args:
            img: Image object from openpyxl
            sheet_name: Name of the sheet
            idx: Image index
            
        Returns:
            Image information dict or None if extraction failed
        """
        try:
            # Get image format
            image_format = img.format if hasattr(img, 'format') else 'unknown'
            
            # Get image data
            image_data = None
            image_filename = None
            
            # Method 1: Try to get from ref (BytesIO object)
            if hasattr(img, 'ref') and img.ref:
                try:
                    ref = img.ref
                    if isinstance(ref, io.BytesIO):
                        image_data = ref.getvalue()
                        logger.debug(f"  Got image data from ref (BytesIO): {len(image_data)} bytes")
                except Exception as e:
                    logger.debug(f"  Failed to get data from ref: {e}")
            
            # Method 2: Try to get from path (if it's a file reference)
            if not image_data and hasattr(img, 'path'):
                path = img.path
                logger.debug(f"  Image path: {path}")
                # Path is usually like /xl/media/image1.jpeg
                # We would need to extract from the XLSX archive
            
            # If we couldn't get image data, return None
            if not image_data:
                logger.warning(f"Could not extract image data for {sheet_name}_{idx:03d}")
                return None
            
            # Determine file extension
            ext = self.SUPPORTED_FORMATS.get(image_format.lower(), '.png')
            image_filename = f"{sheet_name}_{idx:03d}{ext}"
            
            # Save image
            output_path = self.output_dir / image_filename
            with open(output_path, 'wb') as f:
                f.write(image_data)
            
            logger.info(f"  ✓ Saved {image_filename} ({len(image_data)} bytes)")
            
            # Verify image is valid
            try:
                img_obj = Image.open(output_path)
                img_size = img_obj.size
                logger.debug(f"    Image size: {img_size}")
            except Exception as e:
                logger.warning(f"    Warning: Image may be corrupted: {e}")
            
            # Get anchor information
            anchor_info = self._get_anchor_info(img)
            
            return {
                "visual_id": f"{sheet_name}_{idx:03d}",
                "sheet_name": sheet_name,
                "index": idx,
                "filename": image_filename,
                "path": str(output_path),
                "format": image_format,
                "size_bytes": len(image_data),
                "anchor": anchor_info
            }
            
        except Exception as e:
            logger.error(f"Error in _extract_single_image: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _get_anchor_info(self, img) -> Dict[str, Any]:
        """
        Extract anchor information from image.
        
        Args:
            img: Image object
            
        Returns:
            Anchor information dict
        """
        try:
            anchor = img.anchor
            if not anchor:
                return {"status": "no_anchor"}
            
            # Try to get from/to coordinates
            from_row = from_col = to_row = to_col = None
            
            if hasattr(anchor, '_from') and anchor._from:
                from_row = anchor._from.row if hasattr(anchor._from, 'row') else None
                from_col = anchor._from.col if hasattr(anchor._from, 'col') else None
            
            if hasattr(anchor, 'to') and anchor.to:
                to_row = anchor.to.row if hasattr(anchor.to, 'row') else None
                to_col = anchor.to.col if hasattr(anchor.to, 'col') else None
            
            # openpyxl anchor indices are 0-based; convert to 1-based Excel coordinates
            if from_row is not None:
                from_row += 1
            if from_col is not None:
                from_col += 1
            if to_row is not None:
                to_row += 1
            if to_col is not None:
                to_col += 1
            
            if from_row is not None and to_row is None:
                to_row = from_row + 5
            if from_col is not None and to_col is None:
                to_col = from_col + 5
            
            return {
                "from_row": from_row,
                "from_col": from_col,
                "to_row": to_row,
                "to_col": to_col,
                "status": "success"
            }
        except Exception as e:
            logger.debug(f"Error getting anchor info: {e}")
            return {"status": "error", "error": str(e)}
    
    def extract_with_fallback(self, xlsx_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        Extract images with fallback to ZIP extraction if openpyxl fails.
        
        Args:
            xlsx_path: Path to Excel file
            
        Returns:
            Dictionary of extracted images
        """
        # Try primary method
        try:
            return self.extract_from_workbook(xlsx_path)
        except Exception as e:
            logger.warning(f"Primary extraction failed: {e}")
            logger.info("Trying fallback ZIP extraction method...")
            return self._extract_from_zip(xlsx_path)
    
    def _extract_from_zip(self, xlsx_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        Extract images by treating XLSX as ZIP archive.
        
        Args:
            xlsx_path: Path to Excel file
            
        Returns:
            Dictionary of extracted images
        """
        import zipfile
        
        results = {}
        
        try:
            with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
                # Find all media files
                media_files = [f for f in zip_ref.namelist() if 'media/' in f]
                logger.info(f"Found {len(media_files)} media files in archive")
                
                for media_file in media_files:
                    try:
                        # Extract media file
                        media_data = zip_ref.read(media_file)
                        
                        # Generate filename
                        filename = Path(media_file).name
                        output_path = self.output_dir / filename
                        
                        # Save file
                        with open(output_path, 'wb') as f:
                            f.write(media_data)
                        
                        logger.info(f"  ✓ Extracted {filename} ({len(media_data)} bytes)")
                        
                        # Add to results
                        if "unknown" not in results:
                            results["unknown"] = []
                        
                        results["unknown"].append({
                            "visual_id": filename,
                            "filename": filename,
                            "path": str(output_path),
                            "size_bytes": len(media_data),
                            "source": "zip_extraction"
                        })
                    except Exception as e:
                        logger.error(f"Error extracting {media_file}: {e}")
        except Exception as e:
            logger.error(f"ZIP extraction failed: {e}")
        
        return results


def extract_images(xlsx_path: str, output_dir: str = "extracted_images") -> Dict[str, List[Dict[str, Any]]]:
    """
    Convenience function to extract images from Excel file.
    
    Args:
        xlsx_path: Path to Excel file
        output_dir: Output directory
        
    Returns:
        Dictionary of extracted images
    """
    extractor = ImageExtractor(output_dir=Path(output_dir))
    return extractor.extract_with_fallback(xlsx_path)


if __name__ == "__main__":
    # Example usage
    xlsx_file = "RD-03-008-01FMEDAReport.xlsx"
    output_dir = "extracted_images"
    
    extractor = ImageExtractor(output_dir=Path(output_dir))
    results = extractor.extract_with_fallback(xlsx_file)
    
    print("\n" + "=" * 70)
    print("EXTRACTION RESULTS")
    print("=" * 70)
    
    for sheet_name, images in results.items():
        print(f"\nSheet: {sheet_name}")
        print(f"  Total images: {len(images)}")
        for img in images:
            print(f"    - {img['filename']} ({img['size_bytes']} bytes)")
    
    print("\n" + "=" * 70)
    print(f"✓ Extraction completed. Images saved to: {output_dir}")
