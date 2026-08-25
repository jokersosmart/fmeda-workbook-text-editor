"""Acceptance profiling for external FMEDA recalculation results."""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_TARGET_CELLS = (
    "SRAM Tran FIT!T2",
    "SRAM Tran FIT!W2",
    "SRAM Tran FIT!T3",
    "SRAM Tran FIT!W3",
    "SRAM Tran FIT!X2",
    "SRAM Tran FIT!X3",
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _is_error(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("#")


def _read_cells(path: Path, keys: tuple[str, ...]) -> dict[str, dict[str, Any]]:
    formulas = load_workbook(path, data_only=False, read_only=True, keep_links=True)
    cached = load_workbook(path, data_only=True, read_only=True, keep_links=True)
    result: dict[str, dict[str, Any]] = {}
    try:
        for key in keys:
            sheet_name, coordinate = key.split("!", 1)
            if sheet_name not in formulas.sheetnames:
                result[key] = {"present": False, "sheet": sheet_name, "cell": coordinate}
                continue
            formula_cell = formulas[sheet_name][coordinate]
            cached_cell = cached[sheet_name][coordinate]
            formula_raw = (
                formula_cell.value
                if formula_cell.data_type == "f"
                or (isinstance(formula_cell.value, str) and formula_cell.value.startswith("="))
                else None
            )
            cached_value = cached_cell.value
            raw_value = None if formula_raw is not None else formula_cell.value
            result[key] = {
                "present": formula_cell.value is not None or cached_value is not None,
                "sheet": sheet_name,
                "cell": coordinate,
                "formula_raw": formula_raw,
                "raw_value": raw_value,
                "cached_value": cached_value,
                "data_type": formula_cell.data_type,
                "error_state": "error"
                if formula_cell.data_type == "e"
                or _is_error(raw_value)
                or _is_error(cached_value)
                else "ok",
            }
    finally:
        formulas.close()
        cached.close()
    return result


def _formula_equivalent(
    before: str | None,
    after: str | None,
    external_report: dict[str, Any],
) -> bool:
    if before == after:
        return True
    if not before or not after:
        return False
    if external_report.get("mode") != "internal-sheet-copy":
        return False
    normalized = after
    for link in external_report.get("links", []):
        external_sheet = link.get("external_sheet")
        materialized_sheet = link.get("materialized_sheet")
        if external_sheet and materialized_sheet:
            normalized = normalized.replace(
                f"{materialized_sheet}!", f"[{link.get('index', 1)}]{external_sheet}!"
            )
    return normalized == before


def _decision_by_key(manifest: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for decision in (manifest or {}).get("decisions", []):
        if not isinstance(decision, dict):
            continue
        key = str(decision.get("key") or "")
        if key:
            result[key] = decision
    return result


def _external_links(recalc_report: dict[str, Any]) -> list[dict[str, Any]]:
    resolution = recalc_report.get("external_resolution") or {}
    return [item for item in resolution.get("links", []) if isinstance(item, dict)]


class FmedaAcceptanceProfile:
    """Classify external recalculation transitions without silently accepting them."""

    def __init__(self, target_cells: tuple[str, ...] = DEFAULT_TARGET_CELLS):
        self.target_cells = tuple(target_cells)

    def evaluate(
        self,
        base: str | Path,
        target: str | Path,
        recalc_report: dict[str, Any],
        decision_manifest: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        base_path = Path(base).expanduser().resolve()
        target_path = Path(target).expanduser().resolve()
        base_hash = _sha256(base_path)
        base_cells = _read_cells(base_path, self.target_cells)
        target_cells = _read_cells(target_path, self.target_cells)
        external_resolution = recalc_report.get("external_resolution") or {}
        links = _external_links(recalc_report)
        decisions = _decision_by_key(decision_manifest)
        expected_source_hash = recalc_report.get("source_sha256_before")
        provenance_issues: list[dict[str, Any]] = []
        if expected_source_hash and expected_source_hash != base_hash:
            provenance_issues.append(
                {
                    "kind": "source_hash_mismatch",
                    "expected": expected_source_hash,
                    "actual": base_hash,
                    "blocking": True,
                }
            )

        detail: list[dict[str, Any]] = []
        for key in self.target_cells:
            before = base_cells.get(key, {"present": False})
            after = target_cells.get(key, {"present": False})
            reasons: list[str] = []
            evidence: list[str] = []
            status = "review_required"
            if not before.get("present") or not after.get("present"):
                status = "blocked"
                reasons.append("target cell is missing in base or recalculated workbook")
            if before.get("error_state") != "error":
                status = "blocked"
                reasons.append("base cell is not the expected error-state baseline")
            if after.get("error_state") != "ok":
                status = "blocked"
                reasons.append("recalculated cell is still an error")
            if not _formula_equivalent(
                before.get("formula_raw"), after.get("formula_raw"), external_resolution
            ):
                status = "blocked"
                reasons.append("formula is not unchanged or an approved materialization-equivalent")
            if external_resolution.get("status") not in {"BOUND", "MATERIALIZED"}:
                status = "blocked"
                reasons.append("external workbook was not resolved and materialized/bound")
            if not links or any(item.get("status") not in {"resolved", "materialized"} for item in links):
                status = "blocked"
                reasons.append("external link resolution is incomplete")
            if not external_resolution.get("source_kind"):
                status = "blocked"
                reasons.append("external source kind is missing")
            if external_resolution.get("source_kind") == "synthetic-fixture":
                reasons.append("synthetic fixture result cannot be production-accepted")
                evidence.append("synthetic_external_fixture")
            if links and not all(item.get("resolved_sha256") for item in links):
                status = "blocked"
                reasons.append("resolved external workbook hash is missing")
            if provenance_issues:
                status = "blocked"
                reasons.append("source provenance hash does not match")

            decision = decisions.get(key)
            if status == "review_required" and decision:
                decision_status = str(decision.get("status") or "").lower()
                reviewer = str(decision.get("reviewer") or "").strip()
                rationale = str(decision.get("rationale") or "").strip()
                if (
                    decision_status == "accepted"
                    and reviewer
                    and rationale
                    and external_resolution.get("source_kind") != "synthetic-fixture"
                ):
                    status = "accepted"
                    evidence.append(f"reviewer:{reviewer}")
                elif decision_status == "blocked":
                    status = "blocked"
                    reasons.append("reviewer explicitly blocked this result")
                elif decision_status:
                    reasons.append(f"review decision is not sufficient for acceptance: {decision_status}")
            if not reasons:
                reasons.append("external result is numerically available but still requires engineering review")
            detail.append(
                {
                    "key": key,
                    "status": status,
                    "before_error_state": before.get("error_state"),
                    "after_error_state": after.get("error_state"),
                    "before_cached_value": before.get("cached_value"),
                    "after_cached_value": after.get("cached_value"),
                    "formula_before": before.get("formula_raw"),
                    "formula_after": after.get("formula_raw"),
                    "formula_equivalent": _formula_equivalent(
                        before.get("formula_raw"), after.get("formula_raw"), external_resolution
                    ),
                    "reasons": reasons,
                    "evidence": evidence,
                }
            )

        if provenance_issues:
            overall = "blocked"
        elif any(item["status"] == "blocked" for item in detail):
            overall = "blocked"
        elif any(item["status"] == "review_required" for item in detail):
            overall = "review_required"
        else:
            overall = "accepted"
        counts = {status: sum(item["status"] == status for item in detail) for status in (
            "accepted", "review_required", "blocked"
        )}
        return {
            "schema_version": "fmeda-acceptance-report-v1",
            "status": overall,
            "base_file": base_path.name,
            "target_file": target_path.name,
            "base_sha256": base_hash,
            "external_resolution": external_resolution,
            "counts": counts,
            "provenance_issues": provenance_issues,
            "cells": detail,
        }

    @staticmethod
    def render_markdown(report: dict[str, Any]) -> str:
        resolution = report.get("external_resolution") or {}
        counts = report.get("counts") or {}
        lines = [
            "# FMEDA External Recalculation Acceptance Report",
            "",
            f"**Overall status**: `{report['status']}`  ",
            f"**Base**: `{report['base_file']}`  ",
            f"**Target**: `{report['target_file']}`  ",
            f"**External source kind**: `{resolution.get('source_kind', 'unknown')}`  ",
            "",
            "## Manager Summary",
            "",
            "| Status | Count |",
            "|---|---:|",
            f"| Accepted | {counts.get('accepted', 0)} |",
            f"| Review required | {counts.get('review_required', 0)} |",
            f"| Blocked | {counts.get('blocked', 0)} |",
            "",
            "> `review_required` means the calculation is available but no engineering acceptance has been recorded. `blocked` means the evidence is incomplete or contradictory.",
            "",
            "## Reviewer View",
            "",
            "| Location | Status | Before | After | Reason |",
            "|---|---|---|---|---|",
        ]
        for item in report.get("cells", []):
            before = str(item.get("before_cached_value", "<empty>")).replace("|", "\\|")
            after = str(item.get("after_cached_value", "<empty>")).replace("|", "\\|")
            reason = "; ".join(item.get("reasons", []))
            lines.append(f"| `{item['key']}` | `{item['status']}` | `{before}` | `{after}` | {reason} |")
        lines.extend(
            [
                "",
                "## Engineer Evidence",
                "",
                "| Location | Formula equivalent | Formula before | Formula after | Evidence |",
                "|---|---|---|---|---|",
            ]
        )
        for item in report.get("cells", []):
            before = str(item.get("formula_before", "<empty>")).replace("|", "\\|")
            after = str(item.get("formula_after", "<empty>")).replace("|", "\\|")
            evidence = "; ".join(item.get("evidence", [])) or "none"
            lines.append(
                f"| `{item['key']}` | `{item['formula_equivalent']}` | `{before}` | `{after}` | {evidence} |"
            )
        lines.extend(["", "## Provenance", ""])
        if report.get("provenance_issues"):
            for issue in report["provenance_issues"]:
                lines.append(
                    f"- `{issue['kind']}`: expected `{issue['expected']}`, actual `{issue['actual']}`"
                )
        else:
            lines.append(f"- Base SHA-256: `{report['base_sha256']}`")
            lines.append(f"- External resolution: `{resolution.get('status', 'unknown')}`")
            for link in resolution.get("links", []):
                lines.append(
                    f"- External link {link.get('index')}: `{link.get('status')}`, "
                    f"workbook SHA-256 `{link.get('resolved_sha256')}`, "
                    f"materialized sheet `{link.get('materialized_sheet', '<none>')}`"
                )
        lines.extend(
            [
                "",
                "## Decision Rule",
                "",
                "A result is accepted only when the external workbook is non-synthetic, its hash is recorded, the formula is unchanged or materialization-equivalent, the recalculated cell is no longer an error, and a reviewer manifest contains reviewer and rationale.",
                "",
            ]
        )
        return "\n".join(lines)

    def write_reports(
        self,
        base: str | Path,
        target: str | Path,
        recalc_report: dict[str, Any],
        markdown_output: str | Path,
        json_output: str | Path | None = None,
        decision_manifest: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        report = self.evaluate(base, target, recalc_report, decision_manifest)
        markdown_path = Path(markdown_output)
        markdown_path.parent.mkdir(parents=True, exist_ok=True)
        markdown_path.write_text(self.render_markdown(report), encoding="utf-8")
        if json_output:
            json_path = Path(json_output)
            json_path.parent.mkdir(parents=True, exist_ok=True)
            json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        return report
