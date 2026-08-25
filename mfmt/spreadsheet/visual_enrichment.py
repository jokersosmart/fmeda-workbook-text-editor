#!/usr/bin/env python3
"""
visual_enrichment.py - OCR and LLM enrichment for extracted visuals.

Extends Schema A visuals[] with:
- ocr_text
- description (Step 0 summary from LLM analysis)
- analysis_path (separate markdown report file)
- llm_analysis (metadata)
"""

import logging
from pathlib import Path
from typing import Dict, Any, List, Optional

from openpyxl.utils import get_column_letter

from .image_extractor import ImageExtractor

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def _anchor_to_cells(anchor: Dict[str, Any]) -> Dict[str, Any]:
    from_row = anchor.get("from_row")
    from_col = anchor.get("from_col")
    to_row = anchor.get("to_row", from_row)
    to_col = anchor.get("to_col", from_col)
    if from_row is None or from_col is None:
        return {
            "from_cell": None, "from_row": None, "from_col": None,
            "to_cell": None, "to_row": None, "to_col": None,
        }
    # Ensure 1-based indices for get_column_letter (openpyxl anchors are 0-based)
    if from_col < 1:
        from_col += 1
    if to_col is not None and to_col < 1:
        to_col += 1
    if from_row < 1:
        from_row += 1
    if to_row is not None and to_row < 1:
        to_row += 1
    return {
        "from_cell": f"{get_column_letter(from_col)}{from_row}",
        "from_row": from_row,
        "from_col": from_col,
        "to_cell": f"{get_column_letter(to_col)}{to_row}" if to_col and to_row else None,
        "to_row": to_row,
        "to_col": to_col,
    }


class VisualEnricher:
    """Run OCR and optional vision LLM analysis on worksheet visuals."""

    def __init__(
        self,
        images_dir: Path,
        output_base_dir: Optional[Path] = None,
        skip_ocr: bool = False,
        use_llm: bool = False,
        ocr_languages: str = "eng",
        llm_model: Optional[str] = None,
    ):
        self.images_dir = Path(images_dir)
        self.images_dir.mkdir(parents=True, exist_ok=True)
        base = Path(output_base_dir) if output_base_dir else self.images_dir.parent
        self.analysis_dir = base / "analysis"
        self.skip_ocr = skip_ocr
        self.use_llm = use_llm
        self.ocr_languages = ocr_languages
        self.ocr_engine = None
        if not skip_ocr:
            from .image_processor import OCREngine

            self.ocr_engine = OCREngine(max_retries=3)
        self.extractor = ImageExtractor(self.images_dir)
        self.image_analyzer = None
        if use_llm:
            from .image_analysis import ImageContentAnalyzer
            from .llm_config import DEFAULT_VISION_MODEL

            self.image_analyzer = ImageContentAnalyzer(
                model=llm_model or DEFAULT_VISION_MODEL,
                analysis_dir=self.analysis_dir,
            )

    def extract_and_enrich(self, ws, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract images from sheet and enrich with OCR / LLM."""
        visuals: List[Dict[str, Any]] = []
        extracted = self.extractor.extract_from_sheet(ws, sheet_name)

        if not extracted and hasattr(ws, "_images") and ws._images:
            extracted = self._extract_via_blob(ws, sheet_name)

        for img_info in extracted:
            visual = self._build_visual_record(img_info, sheet_name)
            visuals.append(visual)

        return visuals

    def _extract_via_blob(self, ws, sheet_name: str) -> List[Dict[str, Any]]:
        results = []
        for idx, image in enumerate(ws._images, 1):
            try:
                if not hasattr(image, "blob"):
                    continue
                filename = f"{sheet_name}_{idx:03d}.png"
                path = self.images_dir / filename
                with open(path, "wb") as f:
                    f.write(image.blob)
                anchor = {}
                if hasattr(image, "anchor") and image.anchor:
                    a = image.anchor
                    if hasattr(a, "_from"):
                        fr, fc = a._from.row, a._from.col
                        tr = getattr(a._to, "row", fr) if hasattr(a, "_to") else fr
                        tc = getattr(a._to, "col", fc) if hasattr(a, "_to") else fc
                        anchor = {
                            "from_row": fr + 1,
                            "from_col": fc + 1,
                            "to_row": tr + 1,
                            "to_col": tc + 1,
                        }
                results.append({
                    "visual_id": f"{sheet_name}_{idx:03d}",
                    "filename": filename,
                    "path": str(path),
                    "anchor": anchor,
                })
            except Exception as e:
                logger.warning(f"Blob extract failed for {sheet_name} image {idx}: {e}")
        return results

    def _build_visual_record(self, img_info: Dict[str, Any], sheet_name: str) -> Dict[str, Any]:
        visual_id = img_info.get("visual_id") or f"{sheet_name}_{img_info.get('index', 0):03d}"
        filename = img_info.get("filename") or Path(img_info.get("path", "")).name
        image_path = Path(img_info.get("path") or self.images_dir / filename)
        anchor = _anchor_to_cells(img_info.get("anchor", {}))
        image_rel = f"images/{filename}"

        ocr_text = ""
        image_type = "image"
        if not self.skip_ocr and image_path.exists():
            try:
                from .image_processor import ImageClassifier

                classification = ImageClassifier.classify_image(image_path)
                image_type = classification.get("type", "image")
                try:
                    assert self.ocr_engine is not None
                    ocr_text = self.ocr_engine.perform_ocr(image_path, self.ocr_languages)
                except Exception:
                    assert self.ocr_engine is not None
                    ocr_text = self.ocr_engine.perform_ocr(image_path, "eng")
            except Exception as e:
                logger.warning(f"OCR failed for {image_path.name}: {e}")
                ocr_text = "[OCR: Error]"

        visual: Dict[str, Any] = {
            "visual_id": visual_id,
            "type": "image",
            "image_type": image_type,
            "anchor": anchor,
            "image_path": image_rel,
            "ocr_text": ocr_text,
            "description": f"[VISUAL: {visual_id} | {image_type} | {image_rel}]",
        }

        if self.use_llm and self.image_analyzer and image_path.exists():
            result = self.image_analyzer.analyze(
                image_path,
                visual_id=visual_id,
                sheet_name=sheet_name,
                ocr_text=ocr_text,
                cell_anchor=anchor.get("from_cell"),
                image_rel_path=image_rel,
            )
            visual["llm_analysis"] = result
            if result.get("status") == "success":
                if result.get("description"):
                    visual["description"] = result["description"]
                if result.get("analysis_path"):
                    visual["analysis_path"] = result["analysis_path"]

        return visual
