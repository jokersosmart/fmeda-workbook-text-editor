"""Traceable revision diff and validation for FMEDA workspaces."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _is_error(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("#")


def _value_key(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _snapshot(path: Path) -> tuple[dict[str, dict[str, Any]], list[str]]:
    """Read formula/raw values and cached values without modifying the workbook."""
    formulas = load_workbook(path, data_only=False, read_only=True, keep_links=True)
    cached = load_workbook(path, data_only=True, read_only=True, keep_links=True)
    cells: dict[str, dict[str, Any]] = {}
    try:
        sheet_names = list(formulas.sheetnames)
        for sheet_name in sheet_names:
            formula_sheet = formulas[sheet_name]
            cached_sheet = cached[sheet_name]
            for row in formula_sheet.iter_rows():
                for cell in row:
                    raw = cell.value
                    cached_value = cached_sheet[cell.coordinate].value
                    if raw is None and cached_value is None:
                        continue
                    key = f"{sheet_name}!{cell.coordinate}"
                    formula_raw = raw if cell.data_type == "f" or (
                        isinstance(raw, str) and raw.startswith("=")
                    ) else None
                    raw_value = None if formula_raw is not None else raw
                    error_state = (
                        "error"
                        if cell.data_type == "e" or _is_error(raw_value) or _is_error(cached_value)
                        else "ok"
                    )
                    cells[key] = {
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "formula_raw": formula_raw,
                        "raw_value": raw_value,
                        "cached_value": cached_value,
                        "error_state": error_state,
                        "data_type": cell.data_type,
                    }
    finally:
        formulas.close()
        cached.close()
    return cells, sheet_names


def _patch_changes(patch_manifest: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    if not patch_manifest:
        return result
    for change in patch_manifest.get("changes", []):
        if not isinstance(change, dict):
            continue
        sheet = str(change.get("sheet") or "")
        cell = str(change.get("cell") or "").upper().replace("$", "")
        if sheet and cell:
            result[f"{sheet}!{cell}"] = change
    return result


def _display(value: Any) -> str:
    if value is None:
        return "<empty>"
    return str(value).replace("|", "\\|").replace("\n", "\\n")


class FmedaRevisionValidator:
    """Compare two workbooks and validate only explicitly allowed changes."""

    def __init__(
        self,
        base: str | Path,
        target: str | Path,
        patch_manifest: dict[str, Any] | None = None,
    ):
        self.base = Path(base).expanduser().resolve()
        self.target = Path(target).expanduser().resolve()
        self.patch_manifest = patch_manifest or {}
        if not self.base.is_file():
            raise FileNotFoundError(f"base workbook not found: {self.base}")
        if not self.target.is_file():
            raise FileNotFoundError(f"target workbook not found: {self.target}")

    def validate(self) -> dict[str, Any]:
        base_cells, base_sheets = _snapshot(self.base)
        target_cells, target_sheets = _snapshot(self.target)
        allowed = _patch_changes(self.patch_manifest)
        differences: list[dict[str, Any]] = []
        keys = sorted(set(base_cells) | set(target_cells))

        for key in keys:
            before = base_cells.get(key)
            after = target_cells.get(key)
            if before is None or after is None:
                differences.append(
                    {
                        "key": key,
                        "kind": "cell_presence",
                        "blocking": True,
                        "status": "unexpected",
                        "before": before,
                        "after": after,
                    }
                )
                continue
            if before["formula_raw"] != after["formula_raw"]:
                differences.append(
                    {
                        "key": key,
                        "kind": "formula_raw",
                        "blocking": True,
                        "status": "unexpected",
                        "before": before["formula_raw"],
                        "after": after["formula_raw"],
                    }
                )
                continue
            if before["error_state"] != after["error_state"]:
                differences.append(
                    {
                        "key": key,
                        "kind": "error_state",
                        "blocking": True,
                        "status": "unexpected",
                        "before": before["error_state"],
                        "after": after["error_state"],
                    }
                )
            if before["formula_raw"] is not None:
                if _value_key(before["cached_value"]) != _value_key(after["cached_value"]):
                    differences.append(
                        {
                            "key": key,
                            "kind": "cached_value",
                            "blocking": False,
                            "status": "warning",
                            "before": before["cached_value"],
                            "after": after["cached_value"],
                        }
                    )
                continue
            if _value_key(before["raw_value"]) == _value_key(after["raw_value"]):
                continue
            expected = allowed.get(key)
            if expected and _value_key(expected.get("old_value")) == _value_key(before["raw_value"]) and _value_key(
                expected.get("new_value")
            ) == _value_key(after["raw_value"]):
                differences.append(
                    {
                        "key": key,
                        "kind": "allowed_input",
                        "blocking": False,
                        "status": "allowed",
                        "before": before["raw_value"],
                        "after": after["raw_value"],
                    }
                )
            else:
                differences.append(
                    {
                        "key": key,
                        "kind": "unexpected_value",
                        "blocking": True,
                        "status": "unexpected",
                        "before": before["raw_value"],
                        "after": after["raw_value"],
                    }
                )

        for key, expected in allowed.items():
            if key not in base_cells or key not in target_cells:
                continue
            if not any(item["key"] == key and item["kind"] == "allowed_input" for item in differences):
                differences.append(
                    {
                        "key": key,
                        "kind": "missing_expected_change",
                        "blocking": True,
                        "status": "unexpected",
                        "before": base_cells[key].get("raw_value"),
                        "after": target_cells[key].get("raw_value"),
                    }
                )

        if base_sheets != target_sheets:
            differences.append(
                {
                    "key": "workbook",
                    "kind": "sheet_order_or_presence",
                    "blocking": True,
                    "status": "unexpected",
                    "before": base_sheets,
                    "after": target_sheets,
                }
            )

        provenance: list[dict[str, Any]] = []
        base_hash = _sha256(self.base)
        target_hash = _sha256(self.target)
        expected_base_hash = self.patch_manifest.get("base_derived_sha256")
        expected_target_hash = self.patch_manifest.get("derived_sha256")
        if expected_base_hash and expected_base_hash != base_hash:
            provenance.append(
                {
                    "kind": "base_hash_mismatch",
                    "blocking": True,
                    "expected": expected_base_hash,
                    "actual": base_hash,
                }
            )
        if expected_target_hash and expected_target_hash != target_hash:
            provenance.append(
                {
                    "kind": "target_hash_mismatch",
                    "blocking": True,
                    "expected": expected_target_hash,
                    "actual": target_hash,
                }
            )

        blocking = [item for item in differences if item["blocking"]] + [item for item in provenance if item["blocking"]]
        warnings = [item for item in differences if item["status"] == "warning"]
        allowed_count = sum(1 for item in differences if item["status"] == "allowed")
        if blocking:
            status = "FAIL"
        elif warnings:
            status = "PASS_WITH_WARNINGS"
        else:
            status = "PASS"
        return {
            "schema_version": "fmeda-validation-report-v1",
            "status": status,
            "base_file": self.base.name,
            "target_file": self.target.name,
            "base_sha256": base_hash,
            "target_sha256": target_hash,
            "base_sheet_count": len(base_sheets),
            "target_sheet_count": len(target_sheets),
            "allowed_input_changes": allowed_count,
            "blocking_change_count": len(blocking),
            "warning_count": len(warnings),
            "differences": differences,
            "provenance_issues": provenance,
        }

    @staticmethod
    def render_markdown(report: dict[str, Any]) -> str:
        lines = [
            "# FMEDA Revision Validation Report",
            "",
            f"**Status**: `{report['status']}`  ",
            f"**Base**: `{report['base_file']}`  ",
            f"**Target**: `{report['target_file']}`  ",
            f"**Base SHA-256**: `{report['base_sha256']}`  ",
            f"**Target SHA-256**: `{report['target_sha256']}`  ",
            "",
            "## Summary",
            "",
            "| Metric | Count |",
            "|---|---:|",
            f"| Allowed input changes | {report['allowed_input_changes']} |",
            f"| Blocking changes | {report['blocking_change_count']} |",
            f"| Formula cached-value warnings | {report['warning_count']} |",
            "",
            "## Cell and workbook differences",
            "",
            "| Location | Kind | Before | After | Result |",
            "|---|---|---|---|---|",
        ]
        if report["differences"]:
            for item in report["differences"]:
                before = item.get("before")
                after = item.get("after")
                lines.append(
                    f"| `{item['key']}` | `{item['kind']}` | `{_display(before)}` | "
                    f"`{_display(after)}` | `{item['status']}` |"
                )
        else:
            lines.append("| — | none | — | — | `unchanged` |")
        lines.extend(["", "## Provenance issues", ""])
        if report["provenance_issues"]:
            for issue in report["provenance_issues"]:
                lines.append(
                    f"- `{issue['kind']}`: expected `{issue['expected']}`, actual `{issue['actual']}`"
                )
        else:
            lines.append("None.")
        lines.extend(
            [
                "",
                "> 這份報告只驗證 revision 間的可追溯差異，不宣稱已完成 Excel 公式的獨立重算。",
                "",
            ]
        )
        return "\n".join(lines)

    def write_report(self, output: str | Path) -> dict[str, Any]:
        report = self.validate()
        output_path = Path(output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(self.render_markdown(report), encoding="utf-8")
        return report
