"""External workbook discovery and safe temporary binding for FMEDA recalculation."""

from __future__ import annotations

import hashlib
import posixpath
import re
import shutil
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from urllib.parse import unquote, urlparse
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", OFFICE_REL_NS)


class ExternalLinkResolutionError(RuntimeError):
    """Raised when an external link cannot be resolved unambiguously."""


@dataclass(frozen=True)
class ExternalLinkDescriptor:
    index: int
    relationship_id: str
    link_part: str
    original_target: str
    sheet_names: tuple[str, ...]


@dataclass(frozen=True)
class ExternalLinkResolution:
    index: int
    original_target: str
    sheet_names: tuple[str, ...]
    status: str
    resolved_path: str | None = None
    resolved_sha256: str | None = None
    reason: str | None = None


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _relationship_target(source_part: str, target: str) -> str:
    if target.startswith("/") or "://" in target:
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def _parse_relationships(data: bytes) -> dict[str, tuple[str, str, str]]:
    root = ET.fromstring(data)
    result: dict[str, tuple[str, str, str]] = {}
    for relationship in root:
        result[relationship.attrib["Id"]] = (
            relationship.attrib.get("Type", ""),
            relationship.attrib.get("Target", ""),
            relationship.attrib.get("TargetMode", ""),
        )
    return result


def discover_external_links(workbook: str | Path) -> list[ExternalLinkDescriptor]:
    """Read external-link relationships without materializing the workbook."""
    path = Path(workbook).expanduser().resolve()
    with zipfile.ZipFile(path, "r") as archive:
        workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
        workbook_rels = _parse_relationships(archive.read("xl/_rels/workbook.xml.rels"))
        descriptors: list[ExternalLinkDescriptor] = []
        for relationship in workbook_xml:
            if _local(relationship.tag) != "externalReferences":
                continue
            for external_reference in relationship:
                if _local(external_reference.tag) not in {"externalLink", "externalReference"}:
                    continue
                rel_id = external_reference.attrib.get(f"{{{OFFICE_REL_NS}}}id")
                if not rel_id or rel_id not in workbook_rels:
                    continue
                rel_type, rel_target, _ = workbook_rels[rel_id]
                link_part = _relationship_target("xl/workbook.xml", rel_target)
                link_xml = ET.fromstring(archive.read(link_part))
                sheet_names = tuple(
                    node.attrib.get("val", "")
                    for node in link_xml.iter()
                    if _local(node.tag) == "sheetName"
                )
                link_rels_part = posixpath.join(
                    posixpath.dirname(link_part),
                    "_rels",
                    posixpath.basename(link_part) + ".rels",
                )
                link_rels = _parse_relationships(archive.read(link_rels_part))
                external_book = next(
                    (node for node in link_xml.iter() if _local(node.tag) == "externalBook"),
                    None,
                )
                external_rel_id = (
                    external_book.attrib.get(f"{{{OFFICE_REL_NS}}}id")
                    if external_book is not None
                    else None
                )
                if not external_rel_id or external_rel_id not in link_rels:
                    continue
                _, original_target, _ = link_rels[external_rel_id]
                descriptors.append(
                    ExternalLinkDescriptor(
                        index=len(descriptors) + 1,
                        relationship_id=external_rel_id,
                        link_part=link_part,
                        original_target=original_target,
                        sheet_names=sheet_names,
                    )
                )
        return descriptors


def _target_basename(target: str) -> str:
    parsed = urlparse(target)
    raw_path = unquote(parsed.path or target)
    return Path(raw_path.replace("\\", "/")).name.lower()


def resolve_external_links(
    workbook: str | Path,
    external_workbooks: list[str | Path],
) -> list[ExternalLinkResolution]:
    """Resolve each external relationship by an exact basename match."""
    candidates = [
        Path(value).expanduser().resolve()
        for value in external_workbooks
        if Path(value).expanduser().resolve().is_file()
    ]
    descriptors = discover_external_links(workbook)
    resolutions: list[ExternalLinkResolution] = []
    for descriptor in descriptors:
        expected_name = _target_basename(descriptor.original_target)
        matches = [candidate for candidate in candidates if candidate.name.lower() == expected_name]
        if not matches:
            resolutions.append(
                ExternalLinkResolution(
                    index=descriptor.index,
                    original_target=descriptor.original_target,
                    sheet_names=descriptor.sheet_names,
                    status="unresolved",
                    reason=f"no supplied workbook matches basename {expected_name!r}",
                )
            )
            continue
        if len(matches) > 1:
            resolutions.append(
                ExternalLinkResolution(
                    index=descriptor.index,
                    original_target=descriptor.original_target,
                    sheet_names=descriptor.sheet_names,
                    status="ambiguous",
                    reason=f"{len(matches)} supplied workbooks match basename {expected_name!r}",
                )
            )
            continue
        selected = matches[0]
        resolutions.append(
            ExternalLinkResolution(
                index=descriptor.index,
                original_target=descriptor.original_target,
                sheet_names=descriptor.sheet_names,
                status="resolved",
                resolved_path=str(selected),
                resolved_sha256=sha256(selected),
            )
        )
    return resolutions


def _patched_relationship_xml(
    workbook: Path,
    descriptor: ExternalLinkDescriptor,
    resolved_path: Path,
) -> bytes:
    with zipfile.ZipFile(workbook, "r") as archive:
        relationships_part = posixpath.join(
            posixpath.dirname(descriptor.link_part),
            "_rels",
            posixpath.basename(descriptor.link_part) + ".rels",
        )
        root = ET.fromstring(archive.read(relationships_part))
        for relationship in root:
            if relationship.attrib.get("Id") == descriptor.relationship_id:
                relationship.attrib["Target"] = resolved_path.as_uri()
                relationship.attrib["TargetMode"] = "External"
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _patched_external_link_xml(workbook: Path, descriptor: ExternalLinkDescriptor) -> bytes:
    with zipfile.ZipFile(workbook, "r") as archive:
        root = ET.fromstring(archive.read(descriptor.link_part))
    for node in root.iter():
        if _local(node.tag) == "sheetData":
            node.attrib.pop("refreshError", None)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def bind_external_links(
    source: str | Path,
    output: str | Path,
    external_workbooks: list[str | Path],
) -> dict:
    """Create a bound copy whose external relationships point to supplied files.

    Only the output ZIP is changed. The source workbook and external workbooks are
    never written. The formula text remains untouched; the relationship target is
    changed only in the temporary/output copy used for recalculation.
    """
    source_path = Path(source).expanduser().resolve()
    output_path = Path(output).expanduser().resolve()
    resolutions = resolve_external_links(source_path, external_workbooks)
    if not resolutions:
        raise ExternalLinkResolutionError("workbook has no discoverable external links")
    unresolved = [item for item in resolutions if item.status != "resolved"]
    if unresolved:
        detail = "; ".join(f"link {item.index}: {item.status} ({item.reason})" for item in unresolved)
        raise ExternalLinkResolutionError(detail)

    descriptor_by_index = {
        descriptor.index: descriptor
        for descriptor in discover_external_links(source_path)
    }
    patched_relationships: dict[str, bytes] = {}
    patched_links: dict[str, bytes] = {}
    for item in resolutions:
        descriptor = descriptor_by_index[item.index]
        resolved_path = Path(item.resolved_path or "")
        relationships_part = posixpath.join(
            posixpath.dirname(descriptor.link_part),
            "_rels",
            posixpath.basename(descriptor.link_part) + ".rels",
        )
        patched_relationships[relationships_part] = _patched_relationship_xml(
            source_path, descriptor, resolved_path
        )
        patched_links[descriptor.link_part] = _patched_external_link_xml(source_path, descriptor)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(source_path, "r") as source_zip, zipfile.ZipFile(
        output_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as output_zip:
        for info in source_zip.infolist():
            if info.filename in patched_relationships:
                output_zip.writestr(info, patched_relationships[info.filename])
            elif info.filename in patched_links:
                output_zip.writestr(info, patched_links[info.filename])
            else:
                output_zip.writestr(info, source_zip.read(info.filename))

    return {
        "schema_version": "fmeda-external-binding-report-v1",
        "status": "BOUND",
        "source_file": source_path.name,
        "source_sha256": sha256(source_path),
        "bound_file": output_path.name,
        "bound_sha256": sha256(output_path),
        "formula_text_changed": False,
        "links": [asdict(item) for item in resolutions],
    }


def copy_external_workbook(source: str | Path, destination: str | Path) -> None:
    """Copy an external workbook for a disposable calculation workspace."""
    shutil.copy2(Path(source), Path(destination))


def _safe_sheet_name(name: str, used: set[str], prefix: str = "EXT_") -> str:
    candidate = re.sub(r"[^A-Za-z0-9_]", "_", f"{prefix}{name}")[:31] or "EXT_SHEET"
    base = candidate
    suffix = 1
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    return candidate


def _next_relationship_id(root: ET.Element) -> str:
    used = {node.attrib.get("Id", "") for node in root}
    index = 1
    while f"rIdFmedaExt{index}" in used:
        index += 1
    return f"rIdFmedaExt{index}"


def _next_sheet_id(sheets_root: ET.Element) -> str:
    values = []
    for node in sheets_root:
        try:
            values.append(int(node.attrib.get("sheetId", "0")))
        except ValueError:
            continue
    return str(max(values, default=0) + 1)


def _add_materialized_sheet_to_package(
    source: Path,
    output: Path,
    sheet_mappings: list[tuple[str, str, str]],
) -> None:
    """Add internal copies of external sheets and rewrite only external refs."""
    with zipfile.ZipFile(source, "r") as source_zip:
        members = {info.filename: source_zip.read(info.filename) for info in source_zip.infolist()}

    workbook_root = ET.fromstring(members["xl/workbook.xml"])
    workbook_rels_root = ET.fromstring(members["xl/_rels/workbook.xml.rels"])
    sheets_root = next(node for node in workbook_root if _local(node.tag) == "sheets")
    used_sheet_names = {node.attrib.get("name", "") for node in sheets_root}
    content_types_root = ET.fromstring(members["[Content_Types].xml"])
    used_parts = set(members)

    replacements: dict[str, str] = {}
    external_sheet_parts: dict[str, str] = {}
    for index, (external_index, external_sheet, materialized_name) in enumerate(sheet_mappings, start=1):
        external_workbook_path = Path(external_index)
        with zipfile.ZipFile(external_workbook_path, "r") as external_zip:
            external_workbook_root = ET.fromstring(external_zip.read("xl/workbook.xml"))
            external_rels_root = ET.fromstring(external_zip.read("xl/_rels/workbook.xml.rels"))
            external_sheet_node = next(
                node
                for node in external_workbook_root.iter()
                if _local(node.tag) == "sheet" and node.attrib.get("name") == external_sheet
            )
            external_rel_id = external_sheet_node.attrib.get(f"{{{OFFICE_REL_NS}}}id")
            external_rels = _parse_relationships(
                ET.tostring(external_rels_root, encoding="utf-8")
            )
            _, external_sheet_target, _ = external_rels[external_rel_id]
            external_sheet_part = _relationship_target("xl/workbook.xml", external_sheet_target)
            new_sheet_part = f"xl/worksheets/fmeda_external_{index}.xml"
            if new_sheet_part in used_parts:
                raise ExternalLinkResolutionError(
                    f"materialized worksheet part already exists: {new_sheet_part}"
                )
            members[new_sheet_part] = external_zip.read(external_sheet_part)
            used_parts.add(new_sheet_part)

            for content_type in content_types_root:
                if content_type.attrib.get("PartName") == f"/{external_sheet_part}":
                    new_override = dict(content_type.attrib)
                    new_override["PartName"] = f"/{new_sheet_part}"
                    ET.SubElement(content_types_root, content_type.tag, new_override)
                    break
            else:
                ET.SubElement(
                    content_types_root,
                    f"{{{CONTENT_TYPES_NS}}}Override",
                    {
                        "PartName": f"/{new_sheet_part}",
                        "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
                    },
                )

        relation_id = _next_relationship_id(workbook_rels_root)
        ET.SubElement(
            workbook_rels_root,
            f"{{{REL_NS}}}Relationship",
            {
                "Id": relation_id,
                "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
                "Target": f"worksheets/fmeda_external_{index}.xml",
            },
        )
        ET.SubElement(
            sheets_root,
            f"{{{MAIN_NS}}}sheet",
            {
                "name": materialized_name,
                "sheetId": _next_sheet_id(sheets_root),
                f"{{{OFFICE_REL_NS}}}id": relation_id,
            },
        )
        external_sheet_parts[f"{external_sheet}"] = materialized_name
        replacements[f"[{index}]{external_sheet}!"] = f"{materialized_name}!"

    for name, data in list(members.items()):
        if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
            continue
        try:
            root = ET.fromstring(data)
        except ET.ParseError:
            continue
        changed = False
        for node in root.iter():
            if _local(node.tag) != "f" or not node.text:
                continue
            formula = node.text
            formula_changed = False
            for old, new in replacements.items():
                if old in formula:
                    formula = formula.replace(old, new)
                    formula_changed = True
            if formula_changed:
                node.text = formula
                changed = True
        if changed:
            members[name] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    members["xl/workbook.xml"] = ET.tostring(workbook_root, encoding="utf-8", xml_declaration=True)
    members["xl/_rels/workbook.xml.rels"] = ET.tostring(
        workbook_rels_root, encoding="utf-8", xml_declaration=True
    )
    members["[Content_Types].xml"] = ET.tostring(
        content_types_root, encoding="utf-8", xml_declaration=True
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as output_zip:
        for name, data in members.items():
            output_zip.writestr(name, data)


def materialize_external_workbooks(
    source: str | Path,
    output: str | Path,
    external_workbooks: list[str | Path],
) -> dict:
    """Materialize resolved external sheets into a disposable internal-copy workbook.

    The output keeps the source workbook intact and is intended only as a calculation
    input. Formula references such as ``[1]BlockList!`` become ``EXT_BlockList!`` in
    the output copy, while the original external-link package parts remain available
    for provenance inspection.
    """
    source_path = Path(source).expanduser().resolve()
    output_path = Path(output).expanduser().resolve()
    resolutions = resolve_external_links(source_path, external_workbooks)
    if not resolutions:
        raise ExternalLinkResolutionError("workbook has no discoverable external links")
    unresolved = [item for item in resolutions if item.status != "resolved"]
    if unresolved:
        detail = "; ".join(f"link {item.index}: {item.status} ({item.reason})" for item in unresolved)
        raise ExternalLinkResolutionError(detail)

    descriptor_by_index = {
        descriptor.index: descriptor
        for descriptor in discover_external_links(source_path)
    }
    used_names = set(load_workbook(source_path, read_only=True, data_only=False).sheetnames)
    sheet_mappings: list[tuple[str, str, str]] = []
    report_links = []
    for item in resolutions:
        descriptor = descriptor_by_index[item.index]
        for external_sheet in descriptor.sheet_names:
            materialized_name = _safe_sheet_name(external_sheet, used_names)
            used_names.add(materialized_name)
            sheet_mappings.append((item.resolved_path or "", external_sheet, materialized_name))
            report_links.append(
                {
                    **asdict(item),
                    "external_sheet": external_sheet,
                    "materialized_sheet": materialized_name,
                    "status": "materialized",
                }
            )
    _add_materialized_sheet_to_package(source_path, output_path, sheet_mappings)
    return {
        "schema_version": "fmeda-external-materialization-report-v1",
        "status": "MATERIALIZED",
        "mode": "internal-sheet-copy",
        "source_file": source_path.name,
        "source_sha256": sha256(source_path),
        "materialized_file": output_path.name,
        "materialized_sha256": sha256(output_path),
        "formula_text_changed": True,
        "formula_rewrite_scope": "temporary copy only; [n]ExternalSheet! -> EXT_ExternalSheet!",
        "links": report_links,
    }
