#!/usr/bin/env python3
"""
xlsx_to_json.py - Convert Excel (.xlsx) to JSON format with complete metadata preservation.

Optimized for large sheets using efficient iteration and advanced OCR processing.
"""

import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging
from datetime import datetime, date, timezone

import openpyxl
from openpyxl.utils import get_column_letter
from .visual_enrichment import VisualEnricher
from .openpyxl_sparse import effective_dimensions, populated_cells

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelToJsonConverter:
    """Convert Excel workbook to JSON format with metadata preservation."""

    def __init__(self, xlsx_path: str, output_dir: str = "./sheets/", 
                 skip_ocr: bool = False, ocr_languages: str = "eng+chi_tra",
                 max_ocr_retries: int = 3, max_workers: int = 4,
                 use_llm_images: bool = False, llm_model: str = "gpt-4o-mini",
                 extract_images: bool = True):
        self.xlsx_path = Path(xlsx_path)
        self.output_dir = Path(output_dir)
        self.images_dir = self.output_dir / "images"
        self.skip_ocr = skip_ocr
        self.use_llm_images = use_llm_images
        self.extract_images_enabled = extract_images
        self.ocr_languages = ocr_languages
        self.max_ocr_retries = max_ocr_retries
        self.max_workers = max_workers
        self.llm_model = llm_model
        
        self.visual_enricher = VisualEnricher(
            self.images_dir,
            output_base_dir=self.output_dir,
            skip_ocr=skip_ocr,
            use_llm=use_llm_images,
            ocr_languages=ocr_languages,
            llm_model=llm_model,
        )
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.images_dir.mkdir(parents=True, exist_ok=True)
        
        self.wb_formulas = openpyxl.load_workbook(self.xlsx_path, data_only=False)
        self.wb_values = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        
        logger.info(f"Loaded workbook: {self.xlsx_path} ({len(self.wb_formulas.sheetnames)} sheets)")
        logger.info(f"OCR config: languages={ocr_languages}, max_retries={max_ocr_retries}, workers={max_workers}")

    def _serialize_value(self, value: Any) -> Any:
        """Convert value to JSON-serializable format."""
        if value is None:
            return None
        elif isinstance(value, (datetime, date)):
            return value.isoformat()
        elif isinstance(value, bool):
            return value
        elif isinstance(value, (int, float)):
            return value
        else:
            return str(value)

    def extract_formula_references(self, formula: str) -> List[Dict[str, str]]:
        """Parse cross-sheet references from formula."""
        if not formula or not isinstance(formula, str):
            return []
        
        refs = []
        pattern = r"(?:'([^']+)'|([^\s!]+))!(\$?[A-Z]+\$?[0-9]+(?::\$?[A-Z]+\$?[0-9]+)?)"
        
        for match in re.finditer(pattern, formula):
            sheet_name = match.group(1) or match.group(2)
            cell_range = match.group(3)
            refs.append({
                "sheet": sheet_name,
                "range": cell_range,
                "role": "reference"
            })
        
        return refs

    def get_cell_style(self, cell) -> Optional[Dict[str, Any]]:
        """Extract cell style information."""
        style = {}
        
        if cell.font:
            if cell.font.bold:
                style["font_bold"] = True
            if cell.font.size:
                style["font_size"] = cell.font.size
            if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                style["font_color"] = str(cell.font.color.rgb)
        
        if cell.fill and cell.fill.start_color:
            color = cell.fill.start_color
            if hasattr(color, 'rgb') and color.rgb:
                style["bg_color"] = str(color.rgb)
        
        if cell.alignment and cell.alignment.horizontal:
            style["alignment"] = cell.alignment.horizontal
        
        return style if style else None

    def get_data_type(self, cell) -> str:
        """Determine cell data type."""
        if cell.value is None:
            return "null"
        if isinstance(cell.value, bool):
            return "boolean"
        elif isinstance(cell.value, (int, float)):
            return "number"
        elif isinstance(cell.value, str):
            return "error" if cell.value.startswith("#") else "string"
        else:
            return "string"

    def extract_cell_data(self, ws_formulas, ws_values, row: int, col: int) -> Dict[str, Any]:
        """Extract complete cell data."""
        cell_formula = ws_formulas.cell(row, col)
        cell_value = ws_values.cell(row, col)
        return self._extract_cell_data_from_cells(cell_formula, cell_value)

    def _extract_cell_data_from_cells(self, cell_formula, cell_value) -> Dict[str, Any]:
        """Extract one cell without re-looking up the already visited formula cell."""
        row = cell_formula.row
        col = cell_formula.column
        
        col_letter = get_column_letter(col)
        address = f"{col_letter}{row}"
        
        has_formula = cell_formula.value and isinstance(cell_formula.value, str) and cell_formula.value.startswith("=")
        serialized_value = self._serialize_value(cell_value.value)
        
        cell_data = {
            "address": address,
            "value": serialized_value,
            "formula": cell_formula.value if has_formula else None,
            "data_type": self.get_data_type(cell_value),
            "type": "formula" if has_formula else ("empty" if cell_value.value is None else "value"),
        }
        
        if has_formula:
            refs = self.extract_formula_references(cell_formula.value)
            if refs:
                cell_data["formula_refs"] = refs
        
        style = self.get_cell_style(cell_formula)
        if style:
            cell_data["style"] = style
        
        cell_data["merge_anchor"] = None
        
        return cell_data

    def extract_images(self, ws, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract images from worksheet with OCR and optional LLM analysis."""
        return self.visual_enricher.extract_and_enrich(ws, sheet_name)

    def convert_sheet(self, sheet_index: int, sheet_name: str) -> Dict[str, Any]:
        """Convert single sheet to JSON structure."""
        ws_formulas = self.wb_formulas[sheet_name]
        ws_values = self.wb_values[sheet_name]
        
        logger.info(f"Converting sheet {sheet_index}: {sheet_name} (dims: {ws_formulas.dimensions})")
        
        sheet_cells = populated_cells(ws_formulas)
        content_dimensions = effective_dimensions(ws_formulas, sheet_cells)
        sheet_meta = {
            "index": sheet_index,
            "name": sheet_name,
            "dimensions": content_dimensions,
            "declared_dimensions": {
                "max_row": ws_formulas.max_row,
                "max_col": ws_formulas.max_column
            },
            "merged_cells": [str(merged_range) for merged_range in ws_formulas.merged_cells.ranges],
            "tab_color": str(ws_formulas.sheet_properties.tabColor) if ws_formulas.sheet_properties.tabColor else None
        }
        
        cells = {}
        for formula_cell in sheet_cells:
            cell_key = f"{formula_cell.row},{formula_cell.column}"
            value_cell = ws_values.cell(formula_cell.row, formula_cell.column)
            cells[cell_key] = self._extract_cell_data_from_cells(formula_cell, value_cell)
        
        # Extract visual elements with advanced OCR.  Keep compatibility with
        # older test doubles that injected an ``extract_images`` callable
        # directly instead of setting the newer feature flag.
        extract_images_enabled = getattr(self, "extract_images_enabled", None)
        if extract_images_enabled is None:
            extract_images_enabled = callable(getattr(self, "extract_images", None))
        visuals = self.extract_images(ws_formulas, sheet_name) if extract_images_enabled else []
        
        sheet_json = {
            "sheet_meta": sheet_meta,
            "cells": cells,
            "visuals": visuals
        }
        
        return sheet_json

    def convert(self) -> None:
        """Convert entire workbook to JSON files."""
        try:
            self._convert_open_workbooks()
        finally:
            self.wb_formulas.close()
            self.wb_values.close()

    def _convert_open_workbooks(self) -> None:
        """Write the bundle while the formula/value workbook pair is open."""
        logger.info(f"Starting conversion of {self.xlsx_path}")
        sheet_entries: List[Dict[str, Any]] = []

        for sheet_index, sheet_name in enumerate(self.wb_formulas.sheetnames, 1):
            try:
                sheet_json = self.convert_sheet(sheet_index, sheet_name)

                safe_name = re.sub(r'[^\w\s-]', '', sheet_name).strip().replace(' ', '_')
                output_file = self.output_dir / f"{sheet_index:02d}_{safe_name}.json"

                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(sheet_json, f, ensure_ascii=False, indent=2, default=str)

                logger.info(f"✓ Saved: {output_file}")
                sheet_entries.append({
                    "index": sheet_index,
                    "name": sheet_name,
                    "json_file": output_file.name,
                    "md_file": output_file.with_suffix(".md").name,
                })

            except Exception as e:
                logger.error(f"Error converting sheet {sheet_name}: {e}")
                import traceback
                traceback.print_exc()

        manifest = {
            "source_file": str(self.xlsx_path),
            "source_name": Path(self.xlsx_path).name,
            "center_file": f"{Path(self.xlsx_path).stem}_txt.md",
            "exported_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "sheet_count": len(sheet_entries),
            "sheets": sheet_entries,
        }
        manifest_path = self.output_dir / "workbook.json"
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, ensure_ascii=False, indent=2)
        logger.info(f"✓ Saved workbook manifest: {manifest_path}")

        logger.info(f"✓ Conversion complete. Output in: {self.output_dir}")


def main():
    """Command-line entry point."""
    if len(sys.argv) < 2:
        print("Usage: python xlsx_to_json.py <input.xlsx> [--output-dir <dir>] [--skip-ocr] [--llm-images] [--ocr-languages <langs>] [--ocr-retries <n>]")
        sys.exit(1)
    
    xlsx_path = sys.argv[1]
    output_dir = "./sheets/"
    skip_ocr = "--skip-ocr" in sys.argv
    use_llm_images = "--llm-images" in sys.argv
    ocr_languages = "eng+chi_tra"
    max_ocr_retries = 3
    
    if "--output-dir" in sys.argv:
        idx = sys.argv.index("--output-dir")
        if idx + 1 < len(sys.argv):
            output_dir = sys.argv[idx + 1]
    
    if "--ocr-languages" in sys.argv:
        idx = sys.argv.index("--ocr-languages")
        if idx + 1 < len(sys.argv):
            ocr_languages = sys.argv[idx + 1]
    
    if "--ocr-retries" in sys.argv:
        idx = sys.argv.index("--ocr-retries")
        if idx + 1 < len(sys.argv):
            max_ocr_retries = int(sys.argv[idx + 1])
    
    converter = ExcelToJsonConverter(xlsx_path, output_dir, skip_ocr=skip_ocr,
                                    ocr_languages=ocr_languages,
                                    max_ocr_retries=max_ocr_retries,
                                    use_llm_images=use_llm_images)
    converter.convert()


if __name__ == "__main__":
    main()
