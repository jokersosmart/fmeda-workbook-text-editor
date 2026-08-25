from __future__ import annotations

import hashlib
import shutil
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pytest
from openpyxl import Workbook, load_workbook

from mfmt.spreadsheet.fmeda_external import (
    ExternalLinkResolutionError,
    bind_external_links,
    discover_external_links,
)
from mfmt.spreadsheet.fmeda_recalc import LibreOfficeRecalculator

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", OFFICE_REL_NS)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _make_external_book(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BlockList"
    sheet["C2"] = 2
    sheet["C3"] = 4
    sheet["R10"] = "Si MOS: High speed SRAM, FIFO"
    sheet["AO10"] = 4096
    sheet["R11"] = "Si MOS: Digital circuits, Micros, DSP"
    sheet["AO11"] = 8192
    workbook.save(path)


def _make_host_book(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SRAM Tran FIT"
    sheet["U2"] = 172
    sheet["V2"] = 75
    sheet["U3"] = 173
    sheet["V3"] = 306
    sheet["T2"] = '=SUMIF([1]BlockList!R10:R49,"Si MOS: High speed SRAM, FIFO",[1]BlockList!AO10:AO49)'
    sheet["T3"] = '=SUMIF([1]BlockList!R10:R49,"Si MOS: Digital circuits, Micros, DSP",[1]BlockList!AO10:AO49)'
    sheet["W2"] = "=(U2+V2)*T2/1024/1024"
    sheet["W3"] = "=(U3+V3)*T3/1024/1024"
    sheet["X2"] = "=W2/[1]BlockList!C3"
    sheet["X3"] = "=W3/[1]BlockList!C2"
    workbook.save(path)


def _inject_external_link(path: Path, target_name: str) -> None:
    temp = path.with_suffix(".bound.tmp.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp, "w", compression=zipfile.ZIP_DEFLATED
    ) as output:
        members = {info.filename: source.read(info.filename) for info in source.infolist()}
        workbook_root = ET.fromstring(members["xl/workbook.xml"])
        external_refs = next(
            (node for node in workbook_root if node.tag.rsplit("}", 1)[-1] == "externalReferences"),
            None,
        )
        if external_refs is None:
            external_refs = ET.SubElement(workbook_root, f"{{{MAIN_NS}}}externalReferences")
        external_ref = ET.SubElement(
            external_refs,
            f"{{{MAIN_NS}}}externalReference",
            {f"{{{OFFICE_REL_NS}}}id": "rIdExternal1"},
        )
        members["xl/workbook.xml"] = ET.tostring(workbook_root, encoding="utf-8", xml_declaration=True)

        rels_root = ET.fromstring(members["xl/_rels/workbook.xml.rels"])
        ET.SubElement(
            rels_root,
            f"{{{PACKAGE_REL_NS}}}Relationship",
            {
                "Id": "rIdExternal1",
                "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink",
                "Target": "externalLinks/externalLink1.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = ET.tostring(
            rels_root, encoding="utf-8", xml_declaration=True
        )
        members["xl/externalLinks/externalLink1.xml"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<externalLink xmlns="{MAIN_NS}" xmlns:r="{OFFICE_REL_NS}">'
            f'<externalBook r:id="rId1"><sheetNames><sheetName val="BlockList" />'
            f'</sheetNames><sheetDataSet><sheetData sheetId="0" refreshError="1" />'
            f'</sheetDataSet></externalBook></externalLink>'
        ).encode("utf-8")
        members["xl/externalLinks/_rels/externalLink1.xml.rels"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{PACKAGE_REL_NS}">'
            f'<Relationship Id="rId1" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath" '
            f'Target="file:///C:/original/{target_name}" TargetMode="External" />'
            f'</Relationships>'
        ).encode("utf-8")
        for name, data in members.items():
            output.writestr(name, data)
    temp.replace(path)


def test_external_link_is_discovered_and_bound_without_mutating_source(tmp_path: Path) -> None:
    host = tmp_path / "host.xlsx"
    external = tmp_path / "SM2734_HWS_SA_FMEDA_0.2chk.xlsx"
    bound = tmp_path / "bound.xlsx"
    _make_host_book(host)
    _make_external_book(external)
    _inject_external_link(host, external.name)
    source_hash = _sha256(host)

    descriptors = discover_external_links(host)
    assert len(descriptors) == 1
    assert descriptors[0].index == 1
    assert descriptors[0].sheet_names == ("BlockList",)
    assert descriptors[0].original_target.endswith(external.name)

    report = bind_external_links(host, bound, [external])

    assert report["status"] == "BOUND"
    assert report["formula_text_changed"] is False
    assert report["links"][0]["status"] == "resolved"
    assert _sha256(host) == source_hash
    assert bound.is_file()

    formulas = load_workbook(bound, data_only=False)["SRAM Tran FIT"]
    assert formulas["T2"].value.startswith("=SUMIF([1]BlockList!")
    assert formulas["T3"].value.startswith("=SUMIF([1]BlockList!")


def test_external_link_requires_exact_unique_basename(tmp_path: Path) -> None:
    host = tmp_path / "host.xlsx"
    wrong = tmp_path / "other.xlsx"
    _make_host_book(host)
    _make_external_book(wrong)
    _inject_external_link(host, "SM2734_HWS_SA_FMEDA_0.2chk.xlsx")

    with pytest.raises(ExternalLinkResolutionError, match="no supplied workbook"):
        bind_external_links(host, tmp_path / "bound.xlsx", [wrong])


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice is not installed")
def test_materialized_external_workbook_resolves_four_error_state_formulas(tmp_path: Path) -> None:
    host = tmp_path / "host.xlsx"
    external = tmp_path / "SM2734_HWS_SA_FMEDA_0.2chk.xlsx"
    bound = tmp_path / "bound.xlsx"
    recalculated = tmp_path / "recalculated.xlsx"
    _make_host_book(host)
    _make_external_book(external)
    _inject_external_link(host, external.name)
    bind_external_links(host, bound, [external])

    report = LibreOfficeRecalculator(
        bound,
        recalculated,
        timeout_seconds=120,
        external_workbooks=[external],
        external_mode="materialize",
    ).run()

    assert report["external_resolution"]["status"] == "MATERIALIZED"
    assert report["external_resolution"]["links"][0]["materialized_sheet"] == "EXT_BlockList"

    values = load_workbook(recalculated, data_only=True)["SRAM Tran FIT"]
    assert values["T2"].value == 4096
    assert values["T3"].value == 8192
    assert values["W2"].value == pytest.approx((172 + 75) * 4096 / 1024 / 1024)
    assert values["W3"].value == pytest.approx((173 + 306) * 8192 / 1024 / 1024)
