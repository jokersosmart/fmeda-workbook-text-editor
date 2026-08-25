from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from mfmt.spreadsheet.fmeda_workspace import FmedaCoreWorkspaceBuilder


def _make_fixture(path: Path) -> None:
    workbook = Workbook()
    fmeda = workbook.active
    fmeda.title = "FMEDA"
    fmeda["A1"] = "Failure rate input"
    fmeda["B1"] = 0.001
    fmeda["C1"] = "Calculated rate"
    fmeda["D1"] = "=B1*2"
    fmeda["E1"] = "Status"
    fmeda["F1"] = '=IF(B1>0,"OK","REVIEW")'
    fmeda["G1"] = "=[external.xlsx]Missing!A1"
    fmeda.merge_cells("A3:C3")
    fmeda["A3"] = "Merged review heading"

    review = workbook.create_sheet("Review")
    review["A1"] = "Review status"
    review["B1"] = "Pending"
    workbook.save(path)


def test_core_readable_output_has_human_entrypoints_and_machine_links(tmp_path: Path) -> None:
    source = tmp_path / "fmeda.xlsx"
    workspace = tmp_path / "workspace"
    _make_fixture(source)

    manifest = FmedaCoreWorkspaceBuilder(source, workspace).build()

    assert manifest["editor"] is None
    assert manifest["readable"]["index"] == "readable/index.md"
    assert (workspace / "readable" / "index.md").is_file()
    assert (workspace / "readable" / "sheets" / "01_FMEDA.md").is_file()
    assert (workspace / "readable" / "review-queue.md").is_file()
    assert (workspace / "readable" / "formula-guide.md").is_file()

    index = (workspace / "readable" / "index.md").read_text(encoding="utf-8")
    assert "先看這裡" in index
    assert "01_FMEDA.md" in index
    assert "formula_catalog.csv" in index
    assert "dependency_edges.csv" in index
    assert "source_sha256" in index
    assert "優先注意工作表" in index
    assert "FMEDA" in index

    sheet = (workspace / "readable" / "sheets" / "01_FMEDA.md").read_text(
        encoding="utf-8"
    )
    assert "# FMEDA" in sheet
    assert "=B1*2" in sheet
    assert "FMEDA!B1" in sheet
    assert "read_only_formula" in sheet or "cached_value" in sheet
    assert "formula_catalog.csv" in sheet
    assert "merged" in sheet.lower()
    assert "外部引用數" in sheet
    assert "使用的公式函數" in sheet

    review = (workspace / "readable" / "review-queue.md").read_text(encoding="utf-8")
    assert "Review Queue" in review
    assert "FMEDA!G1" in review

    summary = (workspace / "normalized" / "Step03_summary.md").read_text(encoding="utf-8")
    import_report = (workspace / "reports" / "import-report.md").read_text(encoding="utf-8")
    assert "readable/index.md" in summary
    assert "readable/index.md" in import_report


def test_readable_manifest_and_summary_share_core_provenance(tmp_path: Path) -> None:
    source = tmp_path / "fmeda.xlsx"
    workspace = tmp_path / "workspace"
    _make_fixture(source)

    manifest = FmedaCoreWorkspaceBuilder(source, workspace).build()
    readable_manifest = json.loads(
        (workspace / "readable" / "manifest.json").read_text(encoding="utf-8")
    )

    assert readable_manifest["source_sha256"] == manifest["source_sha256"]
    assert readable_manifest["schema_version"] == "fmeda-readable-v1"
    assert readable_manifest["formula_count"] == manifest["formula_count"]
    assert readable_manifest["sheet_count"] == manifest["sheet_count"]


def test_readable_formula_section_is_bounded_for_large_sheets(tmp_path: Path) -> None:
    source = tmp_path / "large.xlsx"
    workspace = tmp_path / "workspace"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FMEDA"
    for row in range(1, 2_201):
        sheet.cell(row, 1, row)
        sheet.cell(row, 2, f"=A{row}*2")
    workbook.save(source)

    FmedaCoreWorkspaceBuilder(source, workspace).build()

    text = (workspace / "readable" / "sheets" / "01_FMEDA.md").read_text(
        encoding="utf-8"
    )
    assert "前 120 筆" in text
    assert text.count("FORMULA-FMEDA-") <= 125


def test_readable_zero_formula_sheet_and_markdown_escaping(tmp_path: Path) -> None:
    source = tmp_path / "plain.xlsx"
    workspace = tmp_path / "workspace"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plain"
    sheet["A1"] = "literal | value\nwith a new line"
    workbook.save(source)

    FmedaCoreWorkspaceBuilder(source, workspace).build()

    text = (workspace / "readable" / "sheets" / "01_Plain.md").read_text(
        encoding="utf-8"
    )
    assert "目前沒有被核心 parser 標記的待審查項目。" in text
    assert "（沒有公式）" in text
    assert "literal \\| value with a new line" in text
