from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import Workbook

from mfmt.spreadsheet.fmeda_acceptance import DEFAULT_TARGET_CELLS, FmedaAcceptanceProfile


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _make_workbook(path: Path, *, error: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SRAM Tran FIT"
    formulas = {
        "T2": '=SUMIF([1]BlockList!R10:R49,"x",[1]BlockList!AO10:AO49)',
        "T3": '=SUMIF([1]BlockList!R10:R49,"y",[1]BlockList!AO10:AO49)',
        "W2": "=(172+75)*T2/1024/1024",
        "W3": "=(173+306)*T3/1024/1024",
        "X2": "=W2/[1]BlockList!C3",
        "X3": "=W3/[1]BlockList!C2",
    }
    for coordinate, formula in formulas.items():
        sheet[coordinate] = formula
    workbook.save(path)
    if not error:
        return
    temp = path.with_suffix(".error.xlsx")
    import zipfile
    from xml.etree import ElementTree as ET

    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp, "w", compression=zipfile.ZIP_DEFLATED
    ) as output:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                root = ET.fromstring(data)
                changed = False
                for cell in root.iter():
                    if cell.tag.rsplit("}", 1)[-1] != "c" or cell.attrib.get("r") not in {
                        "T2", "T3", "W2", "W3", "X2", "X3"
                    }:
                        continue
                    value = next((child for child in cell if child.tag.rsplit("}", 1)[-1] == "v"), None)
                    if value is None:
                        value = ET.SubElement(cell, f"{{{namespace}}}v")
                    cell.attrib["t"] = "e"
                    value.text = "#VALUE!"
                    changed = True
                if changed:
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output.writestr(info, data)
    temp.replace(path)


def _recalc_report(base: Path, *, source_kind: str, status: str = "MATERIALIZED") -> dict:
    return {
        "source_sha256_before": _sha256(base),
        "external_resolution": {
            "status": status,
            "mode": "internal-sheet-copy",
            "source_kind": source_kind,
            "links": [
                {
                    "index": 1,
                    "status": "materialized",
                    "resolved_sha256": "external-sha",
                    "external_sheet": "BlockList",
                    "materialized_sheet": "EXT_BlockList",
                }
            ],
        },
    }


def test_synthetic_result_is_review_required_not_accepted(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    target = tmp_path / "target.xlsx"
    _make_workbook(base, error=True)
    _make_workbook(target, error=False)
    report = FmedaAcceptanceProfile().evaluate(
        base,
        target,
        _recalc_report(base, source_kind="synthetic-fixture"),
    )
    assert report["status"] == "review_required"
    assert report["counts"] == {"accepted": 0, "review_required": 6, "blocked": 0}
    assert all("synthetic fixture" in reason for item in report["cells"] for reason in item["reasons"])


def test_unresolved_external_is_blocked(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    target = tmp_path / "target.xlsx"
    _make_workbook(base, error=True)
    _make_workbook(target, error=False)
    report = FmedaAcceptanceProfile().evaluate(
        base,
        target,
        {"source_sha256_before": _sha256(base), "external_resolution": {"status": "UNRESOLVED_NOT_SUPPLIED", "links": []}},
    )
    assert report["status"] == "blocked"
    assert report["counts"]["blocked"] == 6


def test_production_review_manifest_can_accept_all_cells(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    target = tmp_path / "target.xlsx"
    _make_workbook(base, error=True)
    _make_workbook(target, error=False)
    decisions = {
        "decisions": [
            {
                "key": key,
                "status": "accepted",
                "reviewer": "FMEDA Engineer",
                "rationale": "External BlockList hash and row-level values verified.",
            }
            for key in DEFAULT_TARGET_CELLS
        ]
    }
    report = FmedaAcceptanceProfile().evaluate(
        base,
        target,
        _recalc_report(base, source_kind="production"),
        decisions,
    )
    assert report["status"] == "accepted"
    assert report["counts"] == {"accepted": 6, "review_required": 0, "blocked": 0}
    assert all(item["status"] == "accepted" for item in report["cells"])


def test_formula_change_blocks_even_with_reviewer_manifest(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    target = tmp_path / "target.xlsx"
    _make_workbook(base, error=True)
    _make_workbook(target, error=False)
    workbook = Workbook()
    workbook = __import__("openpyxl").load_workbook(target)
    workbook["SRAM Tran FIT"]["T2"] = "=1"
    workbook.save(target)
    decisions = {
        "decisions": [
            {
                "key": key,
                "status": "accepted",
                "reviewer": "FMEDA Engineer",
                "rationale": "Reviewed.",
            }
            for key in DEFAULT_TARGET_CELLS
        ]
    }
    report = FmedaAcceptanceProfile().evaluate(
        base,
        target,
        _recalc_report(base, source_kind="production"),
        decisions,
    )
    item = next(item for item in report["cells"] if item["key"] == "SRAM Tran FIT!T2")
    assert item["status"] == "blocked"
    assert "formula" in " ".join(item["reasons"])
