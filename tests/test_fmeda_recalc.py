from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from mfmt.spreadsheet.fmeda_recalc import LibreOfficeRecalculator


pytestmark = pytest.mark.skipif(
    shutil.which("soffice") is None,
    reason="LibreOffice is required for recalculation integration test",
)


def test_libreoffice_recalculation_creates_new_file_and_cached_result(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "recalculated.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FMEDA"
    sheet["A1"] = 2
    sheet["B1"] = "=A1*3"
    workbook.save(source)
    source_before = source.read_bytes()

    report = LibreOfficeRecalculator(source, output, timeout_seconds=120).run()

    assert report["status"] == "RECALCULATED"
    assert report["engine"] == "libreoffice-calc"
    assert source.read_bytes() == source_before
    assert output.is_file()
    assert report["formula_count"] == 1
    assert report["cached_result_count"] == 1

    formula_book = load_workbook(output, data_only=False, read_only=True)
    cached_book = load_workbook(output, data_only=True, read_only=True)
    try:
        assert formula_book["FMEDA"]["B1"].value == "=A1*3"
        assert cached_book["FMEDA"]["B1"].value == 6
    finally:
        formula_book.close()
        cached_book.close()
