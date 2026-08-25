from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import Workbook

from mfmt.spreadsheet.fmeda_workspace import FmedaWorkspaceBuilder


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    digest.update(path.read_bytes())
    return digest.hexdigest()


def _make_fixture(path: Path) -> None:
    workbook = Workbook()
    source = workbook.active
    source.title = "Source"
    source["A1"] = 7
    source["B1"] = "=A1+3"
    source["C1"] = "=Missing!A1"

    data = workbook.create_sheet("Data")
    data["A1"] = 11
    data["B1"] = "=Source!B1+Data!A1"
    workbook.save(path)


def test_build_keeps_source_immutable_and_creates_derived_copy(tmp_path: Path) -> None:
    source = tmp_path / "RD-03-008-01FMEDAReport.xlsx"
    workspace = tmp_path / "workspace"
    _make_fixture(source)
    before = _sha256(source)

    result = FmedaWorkspaceBuilder(source, workspace).build()

    assert _sha256(source) == before
    assert result["source_sha256"] == before
    derived = workspace / "derived" / "RD-03-008-01FMEDAReport.rev-001.xlsx"
    assert derived.is_file()
    assert _sha256(derived) == before
    manifest = json.loads((workspace / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_file"] == "source/RD-03-008-01FMEDAReport.xlsx"
    assert manifest["derived_file"] == "derived/RD-03-008-01FMEDAReport.rev-001.xlsx"
    assert manifest["schema_version"] == "workbook-v2"


def test_build_preserves_formula_catalog_and_reference_kinds(tmp_path: Path) -> None:
    source = tmp_path / "fmeda.xlsx"
    workspace = tmp_path / "workspace"
    _make_fixture(source)

    FmedaWorkspaceBuilder(source, workspace).build()

    rows = (workspace / "normalized" / "formula_catalog.csv").read_text(encoding="utf-8")
    assert "Source!B1" in rows
    assert "=A1+3" in rows
    assert "Data!B1" in rows
    assert "=Source!B1+Data!A1" in rows

    edges = (workspace / "normalized" / "dependency_edges.csv").read_text(encoding="utf-8")
    assert ",same_sheet," in edges
    assert ",cross_sheet," in edges
    assert "Source!A1" in edges
    assert "Data!A1" in edges


def test_build_creates_editor_workspace_with_provenance(tmp_path: Path) -> None:
    source = tmp_path / "fmeda.xlsx"
    workspace = tmp_path / "workspace"
    _make_fixture(source)

    FmedaWorkspaceBuilder(source, workspace).build()

    assert (workspace / "normalized" / "Step03_workbook.json").is_file()
    assert (workspace / "normalized" / "Step03_summary.md").is_file()
    assert (workspace / "editor" / "index.md").is_file()
    assert (workspace / "editor" / "sheets" / "01_Source.md").is_file()
    sidecar = json.loads((workspace / "editor" / "blocks.sidecar.json").read_text(encoding="utf-8"))
    assert sidecar["schema_version"] == "fmeda-editor-sidecar-v1"
    assert any(block["source_cell"] == "Source!B1" for block in sidecar["blocks"])
    assert any(block["editability"] == "read_only_formula" for block in sidecar["blocks"])
    relations = json.loads((workspace / "editor" / "relations.json").read_text(encoding="utf-8"))
    assert relations["schema_version"] == "editor-relations-v0.2"


def test_build_keeps_errors_and_external_references_reviewable(tmp_path: Path) -> None:
    source = tmp_path / "fmeda.xlsx"
    workspace = tmp_path / "workspace"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FMEDA"
    sheet["A1"] = "=[External.xlsx]Sheet1!A1"
    sheet["B1"] = "=#DIV/0!"
    workbook.save(source)

    FmedaWorkspaceBuilder(source, workspace).build()

    review = json.loads((workspace / "normalized" / "review_items.json").read_text(encoding="utf-8"))
    kinds = {item["kind"] for item in review["items"]}
    assert "external_reference" in kinds
    assert "formula_error" in kinds
    assert "unresolved" in {item["status"] for item in review["items"] if item["kind"] == "external_reference"}
