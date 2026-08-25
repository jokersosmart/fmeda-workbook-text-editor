from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from mfmt.spreadsheet.fmeda_patch import FmedaPatchApplier
from mfmt.spreadsheet.fmeda_workspace import FmedaWorkspaceBuilder


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _make_source(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FMEDA"
    sheet["A1"] = "Input rate"
    sheet["B1"] = 10
    sheet["C1"] = "Calculated rate"
    sheet["D1"] = "=B1*2"
    sheet["A2"] = "Reviewer note"
    workbook.save(path)


def _build_workspace(tmp_path: Path) -> tuple[Path, dict]:
    source = tmp_path / "fmeda.xlsx"
    workspace = tmp_path / "workspace"
    _make_source(source)
    manifest = FmedaWorkspaceBuilder(source, workspace).build()
    return workspace, manifest


def test_patch_updates_input_only_and_creates_new_revision(tmp_path: Path) -> None:
    workspace, manifest = _build_workspace(tmp_path)
    source = workspace / manifest["source_file"]
    base = workspace / manifest["derived_file"]
    source_hash_before = _sha256(source)
    base_hash_before = _sha256(base)

    patch = {
        "schema_version": "fmeda-patch-v1",
        "patch_id": "review-rate-001",
        "base_source_sha256": manifest["source_sha256"],
        "changes": [
            {
                "sheet": "FMEDA",
                "cell": "B1",
                "editability": "input",
                "expected_old_value": 10,
                "new_value": 15,
            }
        ],
        "review_notes": [
            {
                "source_cell": "FMEDA!B1",
                "author_role": "reviewer",
                "text": "確認輸入 rate 已由審查者更新。",
            }
        ],
    }

    result = FmedaPatchApplier(workspace).apply(patch)

    assert result["status"] == "applied"
    assert result["derived_file"] == "derived/fmeda.rev-002.xlsx"
    assert _sha256(source) == source_hash_before
    assert _sha256(base) == base_hash_before

    revision = workspace / result["derived_file"]
    workbook = load_workbook(revision, data_only=False)
    sheet = workbook["FMEDA"]
    assert sheet["B1"].value == 15
    assert sheet["D1"].value == "=B1*2"
    workbook.close()

    notes = json.loads((workspace / "editor" / "review_notes.json").read_text(encoding="utf-8"))
    assert notes["notes"][0]["patch_id"] == "review-rate-001"
    report = workspace / "reports" / "export-report.rev-002.md"
    assert "FMEDA" in report.read_text(encoding="utf-8")

    with zipfile.ZipFile(revision) as archive:
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
    assert 'calcMode="auto"' in workbook_xml
    assert 'fullCalcOnLoad="1"' in workbook_xml
    assert 'forceFullCalc="1"' in workbook_xml


def test_patch_rejects_formula_cell_and_does_not_create_revision(tmp_path: Path) -> None:
    workspace, manifest = _build_workspace(tmp_path)
    patch = {
        "schema_version": "fmeda-patch-v1",
        "base_source_sha256": manifest["source_sha256"],
        "changes": [
            {
                "sheet": "FMEDA",
                "cell": "D1",
                "editability": "input",
                "new_value": "=B1*4",
            },
        ],
    }

    with pytest.raises(ValueError, match="formula cell is read-only"):
        FmedaPatchApplier(workspace).apply(patch)

    assert not (workspace / "derived" / "fmeda.rev-002.xlsx").exists()


def test_patch_rejects_stale_source_revision(tmp_path: Path) -> None:
    workspace, manifest = _build_workspace(tmp_path)
    patch = {
        "schema_version": "fmeda-patch-v1",
        "base_source_sha256": "stale-source-hash",
        "changes": [
            {
                "sheet": "FMEDA",
                "cell": "B1",
                "editability": "input",
                "new_value": 15,
            },
        ],
    }

    with pytest.raises(ValueError, match="source revision conflict"):
        FmedaPatchApplier(workspace).apply(patch)


def test_patch_rejects_expected_value_conflict(tmp_path: Path) -> None:
    workspace, manifest = _build_workspace(tmp_path)
    patch = {
        "schema_version": "fmeda-patch-v1",
        "base_source_sha256": manifest["source_sha256"],
        "changes": [
            {
                "sheet": "FMEDA",
                "cell": "B1",
                "editability": "input",
                "expected_old_value": 999,
                "new_value": 15,
            }
        ],
    }

    with pytest.raises(ValueError, match="patch conflict"):
        FmedaPatchApplier(workspace).apply(patch)

    assert not (workspace / "derived" / "fmeda.rev-002.xlsx").exists()
