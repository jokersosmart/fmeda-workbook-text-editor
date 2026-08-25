from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from mfmt.spreadsheet.fmeda_large import LargeFmedaValidator


def _make_workbook(
    path: Path,
    *,
    input_value: float = 1.0,
    formula: str = "=B1*2",
    merge: bool = True,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FMEDA"
    sheet["A1"] = "Input"
    sheet["B1"] = input_value
    sheet["C1"] = formula
    sheet["D1"] = "Status"
    sheet["E1"] = "OK"
    if merge:
        sheet.merge_cells("A3:B3")
        sheet["A3"] = "Merged FMEDA section"
    second = workbook.create_sheet("SafetyGoal")
    second["A1"] = "Target"
    second["B1"] = 0.95
    second["C1"] = "=B1"
    workbook.save(path)


def _allowed_patch(base: Path, target: Path) -> dict:
    return {
        "schema_version": "fmeda-patch-manifest-v1",
        "base_derived_sha256": LargeFmedaValidator.sha256(base),
        "derived_sha256": LargeFmedaValidator.sha256(target),
        "changes": [
            {
                "sheet": "FMEDA",
                "cell": "B1",
                "old_value": 1,
                "new_value": 2,
                "editability": "input",
            }
        ],
    }


def test_large_validator_preserves_merges_and_writes_checkpoint(tmp_path: Path) -> None:
    base = tmp_path / "rev-001.xlsx"
    target = tmp_path / "rev-002.xlsx"
    _make_workbook(base, input_value=1.0)
    _make_workbook(target, input_value=2.0)
    output = tmp_path / "validation"

    report = LargeFmedaValidator(base, target, output, _allowed_patch(base, target)).run()

    assert report["status"] == "PASS"
    assert report["summary"]["allowed_input_changes"] == 1
    assert report["summary"]["blocking_changes"] == 0
    assert report["checkpoint"]["status"] == "completed"
    assert report["sheet_summaries"][0]["merged_range_count"] == 1
    assert report["sheet_summaries"][0]["merged_cells_preserved"] is True
    checkpoint = json.loads((output / "checkpoint.json").read_text(encoding="utf-8"))
    assert checkpoint["status"] == "completed"
    assert (output / "chunks" / "01_FMEDA.jsonl").is_file()


def test_large_validator_resumes_after_a_partial_sheet_run(tmp_path: Path) -> None:
    base = tmp_path / "rev-001.xlsx"
    target = tmp_path / "rev-002.xlsx"
    _make_workbook(base, input_value=1.0)
    _make_workbook(target, input_value=2.0)
    output = tmp_path / "validation"
    patch = _allowed_patch(base, target)

    partial = LargeFmedaValidator(base, target, output, patch).run(max_sheets=1)
    assert partial["status"] == "INCOMPLETE"
    assert partial["checkpoint"]["completed_sheet_indices"] == [1]

    resumed = LargeFmedaValidator(base, target, output, patch).run()
    assert resumed["status"] == "PASS"
    assert resumed["checkpoint"]["resumed_sheet_indices"] == [1]
    assert resumed["checkpoint"]["completed_sheet_indices"] == [1, 2]


def test_formula_change_is_blocking_in_streaming_validation(tmp_path: Path) -> None:
    base = tmp_path / "rev-001.xlsx"
    target = tmp_path / "rev-002.xlsx"
    _make_workbook(base, formula="=B1*2")
    _make_workbook(target, formula="=B1*3")

    report = LargeFmedaValidator(base, target, tmp_path / "validation").run()

    assert report["status"] == "FAIL"
    assert any(
        item["kind"] == "formula_raw" and item["blocking"]
        for item in report["samples"]
    )


def test_merged_cell_change_is_blocking(tmp_path: Path) -> None:
    base = tmp_path / "rev-001.xlsx"
    target = tmp_path / "rev-002.xlsx"
    _make_workbook(base, merge=True)
    _make_workbook(target, merge=False)

    report = LargeFmedaValidator(base, target, tmp_path / "validation").run()

    assert report["status"] == "FAIL"
    assert any(
        item["kind"] == "merged_cells" and item["blocking"]
        for item in report["samples"]
    )


def test_equivalent_formula_syntax_is_not_blocking(tmp_path: Path) -> None:
    base = tmp_path / "rev-001.xlsx"
    target = tmp_path / "rev-002.xlsx"
    _make_workbook(base, formula="=VLOOKUP(B1,A1:C2,2,FALSE)")
    _make_workbook(target, formula="= VLOOKUP(B1,A1:C2,2,FALSE())")

    report = LargeFmedaValidator(base, target, tmp_path / "validation").run()

    assert report["status"] == "PASS"
    assert report["summary"]["blocking_changes"] == 0


def test_numeric_recalculation_rounding_is_within_tolerance(tmp_path: Path) -> None:
    base = tmp_path / "rev-001.xlsx"
    target = tmp_path / "rev-002.xlsx"
    _make_workbook(base, input_value=294.2951517732222)
    _make_workbook(target, input_value=294.2951517732220)

    report = LargeFmedaValidator(base, target, tmp_path / "validation").run()

    assert report["status"] == "PASS"
    assert report["summary"]["blocking_changes"] == 0
