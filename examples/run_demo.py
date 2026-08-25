from __future__ import annotations

import hashlib
import json
import shutil
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook

from mfmt.spreadsheet.fmeda_diff import FmedaRevisionValidator
from mfmt.spreadsheet.fmeda_patch import FmedaPatchApplier
from mfmt.spreadsheet.fmeda_workspace import FmedaWorkspaceBuilder
DEMO_ROOT = ROOT / "demo-output"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def make_demo_source(path: Path) -> None:
    workbook = Workbook()
    fmeda = workbook.active
    fmeda.title = "FMEDA"
    fmeda["A1"] = "Failure rate input"
    fmeda["B1"] = 0.001
    fmeda["C1"] = "Calculated rate"
    fmeda["D1"] = "=B1*2"
    fmeda["E1"] = "Status"
    fmeda["F1"] = '=IF(B1>0,"OK","REVIEW")'

    review = workbook.create_sheet("Review")
    review["A1"] = "Review status"
    review["B1"] = "Pending"
    review["C1"] = "This value can be reviewed in Editor."

    safety = workbook.create_sheet("SafetyGoal")
    safety["A1"] = "Target"
    safety["B1"] = 0.95
    safety["C1"] = "Result"
    safety["D1"] = "=B1"
    workbook.save(path)


def main() -> None:
    if DEMO_ROOT.exists():
        shutil.rmtree(DEMO_ROOT)
    source = DEMO_ROOT / "source" / "RD-03-008-01FMEDAReport.xlsx"
    workspace = DEMO_ROOT / "workspace"
    source.parent.mkdir(parents=True, exist_ok=True)
    make_demo_source(source)
    source_before = sha256(source)

    manifest = FmedaWorkspaceBuilder(source, workspace, include_editor=True).build()
    patch = {
        "schema_version": "fmeda-patch-v1",
        "patch_id": "demo-review-rate-001",
        "base_source_sha256": manifest["source_sha256"],
        "changes": [
            {
                "sheet": "FMEDA",
                "cell": "B1",
                "editability": "input",
                "expected_old_value": 0.001,
                "new_value": 0.002,
            }
        ],
        "review_notes": [
            {
                "source_cell": "FMEDA!B1",
                "author_role": "reviewer",
                "text": "審查者確認 failure rate input 已更新。",
            }
        ],
    }
    patch_path = workspace / "normalized" / "demo_patch.json"
    patch_path.write_text(json.dumps(patch, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    result = FmedaPatchApplier(workspace).apply(patch)

    revision = workspace / result["derived_file"]
    validation_report = FmedaRevisionValidator(
        workspace / manifest["derived_file"], revision, result
    ).write_report(workspace / "reports" / "validation.rev-002.md")
    (workspace / "reports" / "validation.rev-002.json").write_text(
        json.dumps(validation_report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    workbook = load_workbook(revision, data_only=False)
    sheet = workbook["FMEDA"]
    input_after = sheet["B1"].value
    formula_after = sheet["D1"].value
    workbook.close()
    source_after = sha256(workspace / manifest["source_file"])

    demo_result = workspace / "DEMO_RESULT.md"
    demo_result.write_text(
        "\n".join(
            [
                "# Slice 2 Demo Result",
                "",
                "> 這個示範展示：Editor／審查 patch 修改 input，公式仍保留，最後產生新的 derived Excel revision。",
                "",
                "| 項目 | 結果 |",
                "|---|---|",
                f"| Original source hash | `{source_before}` |",
                f"| Source hash after patch | `{source_after}` |",
                f"| Patched cell | `FMEDA!B1` |",
                f"| Value before | `0.001` |",
                f"| Value after | `{input_after}` |",
                f"| Formula preserved | `{formula_after}` |",
                f"| Derived revision | `{result['derived_file']}` |",
                f"| Review notes | `{result['review_note_count']}` |",
                f"| Revision validation | `{validation_report['status']}` |",
                "",
                "## 可查看檔案",
                "",
                "- `source/RD-03-008-01FMEDAReport.xlsx`：原始檔，hash 不變。",
                f"- `{result['derived_file']}`：新衍生檔，僅包含受控 input patch。",
                "- `editor/sheets/01_FMEDA.md`：Editor 可讀審查頁。",
                "- `reports/export-report.rev-002.md`：derived export 差異報告。",
                "- `reports/validation.rev-002.md`：全量 revision validation report。",
                "- `editor/review_notes.json`：審查註記。",
                "",
                "公式 cell `FMEDA!D1` 在 Slice 2 中被拒絕修改；Excel workbook.xml 會標記下次開啟時自動重新計算。",
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(demo_result)


if __name__ == "__main__":
    main()
