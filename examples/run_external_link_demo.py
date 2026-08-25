from __future__ import annotations

import json
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from mfmt.spreadsheet.fmeda_recalc import LibreOfficeRecalculator

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", OFFICE_REL_NS)


def local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def make_external_book(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BlockList"
    sheet["C2"] = 2
    sheet["C3"] = 4
    sheet["R10"] = "Si MOS: High speed SRAM, FIFO"
    sheet["AO10"] = 4096
    sheet["R11"] = "Si MOS: Digital circuits, Micros, DSP"
    sheet["AO11"] = 8192
    workbook.save(path)


def make_host_book(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SRAM Tran FIT"
    sheet["U2"] = 172
    sheet["V2"] = 75
    sheet["U3"] = 173
    sheet["V3"] = 306
    sheet["T2"] = '=SUMIF([1]BlockList!R10:R49,"Si MOS: High speed SRAM, FIFO",[1]BlockList!AO10:AO49)'
    sheet["T3"] = '=SUMIF([1]BlockList!R10:R49,"Si MOS: Digital circuits, Micros, DSP",[1]BlockList!AO10:AO49)'
    sheet["W2"] = "=(U2+V2)*T2/1024/1024"
    sheet["W3"] = "=(U3+V3)*T3/1024/1024"
    sheet["X2"] = "=W2/[1]BlockList!C3"
    sheet["X3"] = "=W3/[1]BlockList!C2"
    workbook.save(path)


def inject_external_link(path: Path, target_name: str) -> None:
    temp = path.with_suffix(".external.tmp.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp, "w", compression=zipfile.ZIP_DEFLATED
    ) as output:
        members = {info.filename: source.read(info.filename) for info in source.infolist()}
        workbook_root = ET.fromstring(members["xl/workbook.xml"])
        external_refs = next(
            (node for node in workbook_root if local(node.tag) == "externalReferences"),
            None,
        )
        if external_refs is None:
            external_refs = ET.SubElement(workbook_root, f"{{{MAIN_NS}}}externalReferences")
        ET.SubElement(
            external_refs,
            f"{{{MAIN_NS}}}externalLink",
            {f"{{{OFFICE_REL_NS}}}id": "rIdExternal1"},
        )
        members["xl/workbook.xml"] = ET.tostring(workbook_root, encoding="utf-8", xml_declaration=True)

        rels_root = ET.fromstring(members["xl/_rels/workbook.xml.rels"])
        ET.SubElement(
            rels_root,
            f"{{{PACKAGE_REL_NS}}}Relationship",
            {
                "Id": "rIdExternal1",
                "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink",
                "Target": "externalLinks/externalLink1.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = ET.tostring(
            rels_root, encoding="utf-8", xml_declaration=True
        )
        members["xl/externalLinks/externalLink1.xml"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<externalLink xmlns="{MAIN_NS}" xmlns:r="{OFFICE_REL_NS}">'
            f'<externalBook r:id="rId1"><sheetNames><sheetName val="BlockList" />'
            f'</sheetNames><sheetDataSet><sheetData sheetId="0" refreshError="1" />'
            f'</sheetDataSet></externalBook></externalLink>'
        ).encode("utf-8")
        members["xl/externalLinks/_rels/externalLink1.xml.rels"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{PACKAGE_REL_NS}">'
            f'<Relationship Id="rId1" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath" '
            f'Target="file:///C:/original/{target_name}" TargetMode="External" />'
            f'</Relationships>'
        ).encode("utf-8")
        for name, data in members.items():
            output.writestr(name, data)
    temp.replace(path)


def force_error_cache(path: Path) -> None:
    temp = path.with_suffix(".error.tmp.xlsx")
    target_cells = {"T2", "T3", "W2", "W3"}
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp, "w", compression=zipfile.ZIP_DEFLATED
    ) as output:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                try:
                    root = ET.fromstring(data)
                except ET.ParseError:
                    root = None
                if root is not None:
                    changed = False
                    for cell in root.iter():
                        if local(cell.tag) != "c" or cell.attrib.get("r") not in target_cells:
                            continue
                        value = next((child for child in cell if local(child.tag) == "v"), None)
                        if value is None:
                            value = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
                        cell.attrib["t"] = "e"
                        value.text = "#VALUE!"
                        changed = True
                    if changed:
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output.writestr(info, data)
    temp.replace(path)


def cell_value(path: Path, coordinate: str, data_only: bool) -> object:
    workbook = load_workbook(path, read_only=True, data_only=data_only, keep_links=True)
    value = workbook["SRAM Tran FIT"][coordinate].value
    workbook.close()
    return value


def main() -> int:
    output_dir = ROOT / "demo-output" / "external-link-demo"
    output_dir.mkdir(parents=True, exist_ok=True)
    source = output_dir / "realistic-fmeda-with-external-errors.xlsx"
    external = output_dir / "SM2734_HWS_SA_FMEDA_0.2chk.xlsx"
    materialized = output_dir / "materialized-input.xlsx"
    recalculated = output_dir / "recalculated-with-external-materialized.xlsx"
    report_path = output_dir / "recalculation-report.json"
    summary_path = output_dir / "EXTERNAL_LINK_DEMO.md"

    make_host_book(source)
    inject_external_link(source, external.name)
    force_error_cache(source)
    make_external_book(external)

    report = LibreOfficeRecalculator(
        source,
        recalculated,
        timeout_seconds=120,
        external_workbooks=[external],
        external_mode="materialize",
    ).run()
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    rows = []
    for coordinate in ("T2", "W2", "T3", "W3"):
        rows.append(
            {
                "cell": coordinate,
                "before": cell_value(source, coordinate, data_only=True),
                "after": cell_value(recalculated, coordinate, data_only=True),
            }
        )
    summary = [
        "# External Workbook Materialization Demo",
        "",
        "> This is a synthetic fixture. It proves the loading/materialization path, not the real FMEDA values.",
        "",
        f"- External workbook: `{external.name}`",
        f"- Recalculation mode: `{report['external_resolution']['status']}` / `{report['external_resolution']['mode']}`",
        f"- Materialized sheet: `{report['external_resolution']['links'][0]['materialized_sheet']}`",
        "",
        "| Cell | Before | After |",
        "|---|---:|---:|",
    ]
    summary.extend(f"| `{row['cell']}` | `{row['before']}` | `{row['after']}` |" for row in rows)
    summary.extend(
        [
            "",
            "The source workbook and external fixture are never modified by the recalculation command. The external sheet is copied into a temporary internal sheet so Calc can evaluate the formulas without relying on its `[1]` external-link parser.",
        ]
    )
    summary_path.write_text("\n".join(summary) + "\n", encoding="utf-8")
    print("\n".join(summary))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
